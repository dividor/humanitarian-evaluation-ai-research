#!/usr/bin/env python3
"""
stats.py - Generate organization statistics from Qdrant database

Analyzes document metadata from Qdrant and generates:
- Download/parsing success rates by organization
- Year ranges
- Executive summary detection from metadata
- Abstractive summary statistics
- Sankey diagram visualization
- Excel report with Stats and Stats Viz sheets
"""

import argparse
import logging
import os
import sys
from pathlib import Path

import openpyxl
import plotly.graph_objects as go
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pipeline.db import db

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# TOC completeness threshold - if ratio of pages to TOC items exceeds this,
# the TOC is considered "likely incomplete" and treated as no TOC
TOC_INCOMPLETENESS_RATIO = 15


class StatsGenerator:
    """
    Generate organization statistics and visualizations from Qdrant database.

    This class analyzes processed document metadata to create:
    - Statistical summary by organization (success rates, page counts, etc.)
    - Sankey diagram visualization of pipeline data flow
    - Excel report with embedded visualizations

    Attributes:
        output_path: Path to Excel file for statistics output
    """

    def __init__(self, output_path="./data/pipeline_stats.xlsx"):
        """
        Initialize the StatsGenerator.

        Args:
            output_path: Path to output Excel file for statistics
        """
        self.output_path = output_path

    def has_executive_summary(self, key_content_sections):
        """
        Check if document has executive summary from Key Content Sections metadata.

        Examines the Key Content Sections field to determine if the document
        contains identified summary sections (Executive Summary, Abstract, etc.).

        Args:
            key_content_sections: Value from 'key_content_sections' field in Qdrant

        Returns:
            bool: True if has summary sections, False otherwise
        """
        if not key_content_sections or not isinstance(key_content_sections, str):
            return False

        # Check for specific values that indicate no summary
        if key_content_sections in ["No", "No TOC", ""]:
            return False

        # If it has content (TOC entries), it has summary sections
        return len(key_content_sections.strip()) > 0

    def load_data_from_qdrant(self):
        """
        Load all document metadata from Qdrant database.

        Returns:
            list: List of document dictionaries with metadata
        """
        logger.info("Loading documents from Qdrant...")
        documents = list(db.get_all_documents())
        logger.info("✓ Loaded %d documents from Qdrant", len(documents))
        return documents

    def calculate_stats(self, documents):
        """
        Calculate statistics by organization from document list.

        Iterates through all documents, aggregating metrics for each organization:
        - Download/parse success counts
        - TOC and summary presence
        - Page and word counts (totals and averages)
        - Year ranges

        Args:
            documents: List of document dictionaries from Qdrant

        Returns:
            dict: Nested dictionary with organization as key, containing statistics:
                {
                    'org_name': {
                        'years': [list of years],
                        'total': int,
                        'downloaded': int,
                        'parsed': int,
                        'has_toc': int,
                        'has_exec_summary': int,
                        'has_abstractive_summary': int,
                        'total_pages': int,
                        'total_words': int,
                        'page_count': int,
                        'word_count': int
                    }
                }
        """
        logger.info("Calculating statistics by organization...")

        # Organize data by agency
        agency_data = {}

        for doc in documents:
            agency = doc.get("organization") or doc.get("agency")
            if not agency:
                continue

            if agency not in agency_data:
                agency_data[agency] = {
                    "years": [],
                    "total": 0,
                    "downloaded": 0,
                    "parsed": 0,
                    "has_toc": 0,
                    "has_exec_summary": 0,
                    "has_abstractive_summary": 0,
                    "total_pages": 0,
                    "total_words": 0,
                    "page_count": 0,  # Count of docs with page data
                    "word_count": 0,  # Count of docs with word data
                }

            data = agency_data[agency]
            data["total"] += 1

            # Year
            year = doc.get("year")
            if year:
                data["years"].append(year)

            # Download success (check status or error field)
            status = doc.get("status", "")
            download_error = doc.get("download_error") or doc.get("error")
            if status in ["downloaded", "parsed", "summarized"] or (
                not download_error and doc.get("filepath")
            ):
                data["downloaded"] += 1

            # Parse success
            if status in ["parsed", "summarized"] or doc.get("parsed_folder"):
                data["parsed"] += 1

            # Has TOC
            toc = doc.get("toc")
            if toc and isinstance(toc, str) and len(toc.strip()) > 0:
                data["has_toc"] += 1

            # Has Executive Summary (from key_content_sections metadata)
            key_content_sections = doc.get("key_content_sections")
            if self.has_executive_summary(key_content_sections):
                data["has_exec_summary"] += 1

            # Has Abstractive Summary
            abstractive_summary = doc.get("full_summary")
            if (
                abstractive_summary
                and isinstance(abstractive_summary, str)
                and len(abstractive_summary.strip()) > 0
            ):
                data["has_abstractive_summary"] += 1

            # Track page and word counts
            page_count = doc.get("page_count")
            if page_count and isinstance(page_count, (int, float)) and page_count > 0:
                data["total_pages"] += page_count
                data["page_count"] += 1

            word_count = doc.get("word_count")
            if word_count and isinstance(word_count, (int, float)) and word_count > 0:
                data["total_words"] += word_count
                data["word_count"] += 1

        return agency_data

    def create_stats_sheet(self, wb, agency_data):
        """
        Create or update Stats sheet with organization statistics.

        Generates a formatted Excel sheet with per-organization metrics including:
        - Year ranges
        - Success rates and counts (download, parse, TOC, summaries)
        - Average page and word counts

        Replaces any existing Stats sheet and applies header styling.

        Args:
            wb: openpyxl Workbook object
            agency_data: Dictionary of statistics by organization (from calculate_stats)
        """
        logger.info("Creating Stats sheet...")

        # Remove existing Stats sheet if it exists
        if "Stats" in wb.sheetnames:
            del wb["Stats"]

        # Create new Stats sheet at the end
        stats_ws = wb.create_sheet("Stats")
        logger.info("Stats sheet will be placed to the right of existing tabs")

        # Define headers
        headers = [
            "Organization",
            "Start Year",
            "End Year",
            "Total Number of PDFs",
            "% PDFs Downloaded Successfully",
            "Number of PDFs Downloaded Successfully",
            "% PDFs Parsed Successfully",
            "Number of PDFs Parsed Successfully",
            "% PDFs with Table of Contents",
            "Number of PDFs with Table of Contents",
            "% PDFs with Summary Section",
            "Number of PDFs with Summary Section",
            "% PDFs with Abstractive Summary",
            "Number of PDFs with Abstractive Summary",
            "Average Number of Pages per PDF",
            "Average Number of Words per PDF",
        ]

        # Write headers with styling
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = stats_ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font

        # Sort agencies alphabetically
        sorted_agencies = sorted(agency_data.keys())

        # Write data rows
        for row_idx, agency in enumerate(sorted_agencies, 2):
            data = agency_data[agency]

            # Calculate percentages
            pct_downloaded = (
                (data["downloaded"] / data["total"] * 100) if data["total"] > 0 else 0
            )
            pct_parsed = (
                (data["parsed"] / data["total"] * 100) if data["total"] > 0 else 0
            )
            pct_toc = (
                (data["has_toc"] / data["total"] * 100) if data["total"] > 0 else 0
            )
            pct_exec_summary = (
                (data["has_exec_summary"] / data["total"] * 100)
                if data["total"] > 0
                else 0
            )
            pct_abstractive = (
                (data["has_abstractive_summary"] / data["total"] * 100)
                if data["total"] > 0
                else 0
            )

            # Calculate averages
            avg_pages = (
                (data["total_pages"] / data["page_count"])
                if data["page_count"] > 0
                else 0
            )
            avg_words = (
                (data["total_words"] / data["word_count"])
                if data["word_count"] > 0
                else 0
            )

            # Year range
            start_year = min(data["years"]) if data["years"] else ""
            end_year = max(data["years"]) if data["years"] else ""

            # Write row
            row_data = [
                agency,
                start_year,
                end_year,
                data["total"],
                round(pct_downloaded, 1),
                data["downloaded"],
                round(pct_parsed, 1),
                data["parsed"],
                round(pct_toc, 1),
                data["has_toc"],
                round(pct_exec_summary, 1),
                data["has_exec_summary"],
                round(pct_abstractive, 1),
                data["has_abstractive_summary"],
                round(avg_pages, 1),
                round(avg_words, 1),
            ]

            for col_idx, value in enumerate(row_data, 1):
                stats_ws.cell(row_idx, col_idx, value)

        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            stats_ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(len(header) + 2, 15)

        logger.info("✓ Stats sheet created with %d organizations", len(sorted_agencies))

    def create_stats_viz_sheet(self, wb, image_path="./data/pipeline_sankey.png"):
        """
        Create or update Stats Viz sheet with embedded Sankey diagram.

        Embeds the pipeline flow visualization image into a dedicated Excel sheet
        for easy viewing alongside the statistical tables.

        Args:
            wb: openpyxl Workbook object
            image_path: Path to the Sankey diagram PNG file to embed
        """
        logger.info("Creating Stats Viz sheet...")

        # Remove existing Stats Viz sheet if it exists
        if "Stats Viz" in wb.sheetnames:
            del wb["Stats Viz"]

        # Create new Stats Viz sheet
        viz_ws = wb.create_sheet("Stats Viz")
        logger.info("Stats Viz sheet will be placed to the right of existing tabs")

        # Check if image file exists
        if not Path(image_path).exists():
            logger.warning("Sankey diagram not found at: %s", image_path)
            logger.warning("Skipping image embedding")
            return

        # Embed the image
        try:
            img = Image(image_path)
            # Position the image at cell A1
            viz_ws.add_image(img, "A1")
            logger.info("✓ Sankey diagram embedded in Stats Viz sheet")
        except Exception as e:
            logger.error("Failed to embed image: %s", e)

    def create_sankey_diagram(
        self, documents, output_path="./data/pipeline_sankey.png"
    ):
        """
        Create Sankey diagram showing data flow through pipeline stages.

        Generates an interactive Sankey visualization that shows how documents
        flow through the pipeline:
        1. Organization → Downloaded/Not Downloaded
        2. Downloaded → Parsed/Not Parsed
        3. Parsed → Has TOC/No TOC
        4. Has TOC/No TOC → Has Summary/No Summary

        Applies TOC completeness threshold to filter out incomplete TOCs.
        Exports as high-resolution PNG for embedding in Excel and standalone viewing.

        Args:
            documents: List of document dictionaries from Qdrant
            output_path: Path to save the PNG file (default: './data/pipeline_sankey.png')
        """
        logger.info("Creating Sankey diagram...")

        # Track flows by organization
        org_flows = {}
        all_years = []

        # Process each document
        for doc in documents:
            agency = doc.get("organization") or doc.get("agency")
            if not agency:
                continue

            # Collect year data
            year = doc.get("year")
            if year:
                all_years.append(year)

            if agency not in org_flows:
                org_flows[agency] = {
                    "total": 0,
                    "downloaded": 0,
                    "not_downloaded": 0,
                    "parsed": 0,
                    "not_parsed": 0,
                    "has_toc": 0,
                    "no_toc": 0,
                    "has_summary_with_toc": 0,
                    "no_summary_with_toc": 0,
                    "has_summary_no_toc": 0,
                    "no_summary_no_toc": 0,
                }

            flows = org_flows[agency]
            flows["total"] += 1

            # Track download status
            status = doc.get("status", "")
            download_error = doc.get("download_error") or doc.get("error")
            downloaded = False
            if status in ["downloaded", "parsed", "summarized"] or (
                not download_error and doc.get("filepath")
            ):
                flows["downloaded"] += 1
                downloaded = True
            else:
                flows["not_downloaded"] += 1

            # Track parse status (only for downloaded)
            parsed = False
            if downloaded:
                if status in ["parsed", "summarized"] or doc.get("parsed_folder"):
                    flows["parsed"] += 1
                    parsed = True
                else:
                    flows["not_parsed"] += 1

            # Track TOC status (only for parsed)
            has_toc = False
            if parsed:
                toc = doc.get("toc")
                if toc and isinstance(toc, str) and len(toc.strip()) > 0:
                    # Count TOC entries
                    toc_lines = toc.split("\n")
                    toc_entry_count = 0
                    for line in toc_lines:
                        line = line.strip()
                        if (
                            line
                            and not line.startswith("---")
                            and "[H" in line
                            and "|" in line
                        ):
                            toc_entry_count += 1

                    # Get page count
                    page_count = doc.get("page_count")
                    if page_count and isinstance(page_count, (int, float)):
                        page_count = int(page_count)

                    # Check if TOC is likely incomplete (ratio of pages to TOC items > threshold)
                    if page_count and toc_entry_count > 0:
                        ratio = page_count / toc_entry_count
                        if ratio > TOC_INCOMPLETENESS_RATIO:
                            # TOC exists but is likely incomplete
                            flows["no_toc"] += 1
                        else:
                            flows["has_toc"] += 1
                            has_toc = True
                    else:
                        # Can't calculate ratio, just count as has_toc
                        flows["has_toc"] += 1
                        has_toc = True
                else:
                    flows["no_toc"] += 1

            # Track summary status (for both TOC and non-TOC parsed docs)
            if parsed:
                abstractive_summary = doc.get("full_summary")
                has_summary = (
                    abstractive_summary
                    and isinstance(abstractive_summary, str)
                    and len(abstractive_summary.strip()) > 0
                )

                if has_toc:
                    if has_summary:
                        flows["has_summary_with_toc"] += 1
                    else:
                        flows["no_summary_with_toc"] += 1
                else:
                    # Also track summaries for documents without TOC
                    if has_summary:
                        flows["has_summary_no_toc"] += 1
                    else:
                        flows["no_summary_no_toc"] += 1

        # Build Sankey diagram
        # Define modern muted color palette for organizations
        org_colors = [
            "rgb(102, 194, 165)",  # Muted teal
            "rgb(252, 141, 98)",  # Muted coral
            "rgb(141, 160, 203)",  # Muted blue
            "rgb(231, 138, 195)",  # Muted pink
            "rgb(166, 216, 84)",  # Muted green
            "rgb(255, 217, 47)",  # Muted yellow
            "rgb(229, 196, 148)",  # Muted tan
            "rgb(179, 179, 179)",  # Muted gray
        ]

        # Create node labels and map organizations to colors
        nodes = []
        node_colors = []
        node_idx = {}
        org_color_map = {}

        sorted_orgs = sorted(org_flows.keys())
        for idx, org in enumerate(sorted_orgs):
            org_color_map[org] = org_colors[idx % len(org_colors)]

        # Layer 1: Total by org (use <br> for line breaks)
        for org in sorted_orgs:
            node_idx[f"{org}_total"] = len(nodes)
            # For small nodes, put count inline to avoid overlap
            nodes.append(f"{org} (n={org_flows[org]['total']})")
            node_colors.append(org_color_map[org])

        # Layer 2: Downloaded / Not Downloaded
        node_idx["downloaded"] = len(nodes)
        nodes.append("Downloaded")
        node_colors.append("rgba(144, 238, 144, 0.8)")

        node_idx["not_downloaded"] = len(nodes)
        nodes.append("Problems<br>Downloading")
        node_colors.append("rgba(255, 99, 71, 0.8)")

        # Layer 3: Parsed / Not Parsed
        node_idx["parsed"] = len(nodes)
        nodes.append("Parsed")
        node_colors.append("rgba(135, 206, 250, 0.8)")

        node_idx["not_parsed"] = len(nodes)
        nodes.append("Not Parsed")
        node_colors.append("rgba(255, 160, 122, 0.8)")

        # Layer 4: Has TOC / No TOC
        node_idx["has_toc"] = len(nodes)
        nodes.append("Report Headings<br>Extracted")
        node_colors.append("rgba(147, 112, 219, 0.8)")

        node_idx["no_toc"] = len(nodes)
        nodes.append("Problems Parsing<br>Document Structure")
        node_colors.append("rgba(240, 128, 128, 0.8)")

        # Layer 5: Has Summary / No Summary
        node_idx["has_summary"] = len(nodes)
        nodes.append("Has AI<br>Summary")
        node_colors.append("rgba(60, 179, 113, 0.8)")

        node_idx["no_summary"] = len(nodes)
        nodes.append("No AI<br>Summary")
        node_colors.append("rgba(205, 92, 92, 0.8)")

        # Build links
        sources = []
        targets = []
        values = []
        link_colors = []

        # Layer 1 -> Layer 2 (Org to Downloaded/Not Downloaded)
        for org in sorted_orgs:
            flows = org_flows[org]
            if flows["downloaded"] > 0:
                sources.append(node_idx[f"{org}_total"])
                targets.append(node_idx["downloaded"])
                values.append(flows["downloaded"])
                link_colors.append(
                    org_color_map[org].replace("rgb", "rgba").replace(")", ", 0.4)")
                )

            if flows["not_downloaded"] > 0:
                sources.append(node_idx[f"{org}_total"])
                targets.append(node_idx["not_downloaded"])
                values.append(flows["not_downloaded"])
                link_colors.append(
                    org_color_map[org].replace("rgb", "rgba").replace(")", ", 0.4)")
                )

        # Layer 2 -> Layer 3 (Downloaded to Parsed/Not Parsed)
        total_downloaded_parsed = sum(org_flows[org]["parsed"] for org in sorted_orgs)
        total_downloaded_not_parsed = sum(
            org_flows[org]["not_parsed"] for org in sorted_orgs
        )

        if total_downloaded_parsed > 0:
            sources.append(node_idx["downloaded"])
            targets.append(node_idx["parsed"])
            values.append(total_downloaded_parsed)
            link_colors.append("rgba(144, 238, 144, 0.4)")

        if total_downloaded_not_parsed > 0:
            sources.append(node_idx["downloaded"])
            targets.append(node_idx["not_parsed"])
            values.append(total_downloaded_not_parsed)
            link_colors.append("rgba(255, 99, 71, 0.4)")

        # Layer 3 -> Layer 4 (Parsed to Has TOC/No TOC)
        total_parsed_toc = sum(org_flows[org]["has_toc"] for org in sorted_orgs)
        total_parsed_no_toc = sum(org_flows[org]["no_toc"] for org in sorted_orgs)

        if total_parsed_toc > 0:
            sources.append(node_idx["parsed"])
            targets.append(node_idx["has_toc"])
            values.append(total_parsed_toc)
            link_colors.append("rgba(135, 206, 250, 0.4)")

        if total_parsed_no_toc > 0:
            sources.append(node_idx["parsed"])
            targets.append(node_idx["no_toc"])
            values.append(total_parsed_no_toc)
            link_colors.append("rgba(255, 160, 122, 0.4)")

        # Layer 4 -> Layer 5 (Has TOC to Has Summary/No Summary)
        total_toc_summary = sum(
            org_flows[org]["has_summary_with_toc"] for org in sorted_orgs
        )
        total_toc_no_summary = sum(
            org_flows[org]["no_summary_with_toc"] for org in sorted_orgs
        )

        if total_toc_summary > 0:
            sources.append(node_idx["has_toc"])
            targets.append(node_idx["has_summary"])
            values.append(total_toc_summary)
            link_colors.append("rgba(147, 112, 219, 0.4)")

        if total_toc_no_summary > 0:
            sources.append(node_idx["has_toc"])
            targets.append(node_idx["no_summary"])
            values.append(total_toc_no_summary)
            link_colors.append("rgba(240, 128, 128, 0.4)")

        # Layer 4 -> Layer 5 (No TOC to Has Summary/No Summary)
        total_no_toc_summary = sum(
            org_flows[org]["has_summary_no_toc"] for org in sorted_orgs
        )
        total_no_toc_no_summary = sum(
            org_flows[org]["no_summary_no_toc"] for org in sorted_orgs
        )

        if total_no_toc_summary > 0:
            sources.append(node_idx["no_toc"])
            targets.append(node_idx["has_summary"])
            values.append(total_no_toc_summary)
            link_colors.append("rgba(240, 128, 128, 0.4)")

        if total_no_toc_no_summary > 0:
            sources.append(node_idx["no_toc"])
            targets.append(node_idx["no_summary"])
            values.append(total_no_toc_no_summary)
            link_colors.append("rgba(240, 128, 128, 0.4)")

        # Create Sankey diagram with modern styling
        fig = go.Figure(
            data=[
                go.Sankey(
                    node=dict(
                        pad=35,  # Increased padding to give more space for small nodes
                        thickness=20,
                        line=dict(
                            color="white", width=0
                        ),  # No borders for cleaner look
                        label=nodes,
                        color=node_colors,
                    ),
                    link=dict(
                        source=sources,
                        target=targets,
                        value=values,
                        color=link_colors,
                        line=dict(
                            color="rgba(0,0,0,0.1)", width=0.5
                        ),  # Light outline for visibility
                        hovertemplate="<b>%{source.label}</b> → <b>%{target.label}</b><br>"
                        + "Reports: %{value:,}<br>"
                        + "<extra></extra>",
                    ),
                    textfont=dict(size=13),
                    arrangement="snap",  # Better automatic layout
                    valueformat=",",  # Add commas to numbers
                )
            ]
        )

        # Calculate totals for layer labels
        total_records = sum(org_flows[org]["total"] for org in sorted_orgs)
        num_orgs = len(sorted_orgs)
        total_layer2 = (
            total_downloaded_parsed
            + total_downloaded_not_parsed
            + sum(org_flows[org]["not_downloaded"] for org in sorted_orgs)
        )
        total_layer3 = total_parsed_toc + total_parsed_no_toc
        total_layer4 = (
            total_toc_summary
            + total_toc_no_summary
            + total_no_toc_summary
            + total_no_toc_no_summary
        )

        # Calculate year range
        year_range = ""
        if all_years:
            min_year = min(all_years)
            max_year = max(all_years)
            year_range = f" (Data from {min_year} to {max_year})"

        fig.update_layout(
            title_text=(
                f"<b>UN Humanitarian Evaluation Reports Processing Pipeline "
                f"Results{year_range}</b>"
            ),
            title_x=0.5,  # Center align title
            title_font_size=24,  # Double the default size (was ~12, now 24)
            font_size=12,
            height=1000,
            width=1400,
            annotations=[
                # Layer 1 label (below)
                dict(
                    x=0.02,
                    y=-0.08,
                    xref="paper",
                    yref="paper",
                    text=f"UN Organization (orgs={num_orgs:,}, reports={total_records:,})",
                    showarrow=False,
                    font=dict(size=14, color="black"),
                    xanchor="left",
                ),
                # Layer 2 label (above the nodes) - white pill background
                dict(
                    x=0.28,
                    y=0.97,
                    xref="paper",
                    yref="paper",
                    text=f"reports={total_layer2:,}",
                    showarrow=False,
                    font=dict(size=14, color="black"),
                    xanchor="center",
                    bgcolor="white",
                    borderpad=8,
                    bordercolor="rgba(200,200,200,0.5)",
                    borderwidth=1,
                ),
                # Layer 3 label (above the nodes) - white pill background
                dict(
                    x=0.51,
                    y=0.97,
                    xref="paper",
                    yref="paper",
                    text=f"reports={total_layer3:,}",
                    showarrow=False,
                    font=dict(size=14, color="black"),
                    xanchor="center",
                    bgcolor="white",
                    borderpad=8,
                    bordercolor="rgba(200,200,200,0.5)",
                    borderwidth=1,
                ),
                # Layer 4 label (above the nodes) - white pill background
                dict(
                    x=0.74,
                    y=0.97,
                    xref="paper",
                    yref="paper",
                    text=f"reports={total_layer4:,}",
                    showarrow=False,
                    font=dict(size=14, color="black"),
                    xanchor="center",
                    bgcolor="white",
                    borderpad=8,
                    bordercolor="rgba(200,200,200,0.5)",
                    borderwidth=1,
                ),
            ],
        )

        # Save as PNG with high resolution
        logger.info("Saving Sankey diagram to: %s", output_path)
        fig.write_image(output_path, scale=3)  # High resolution export
        logger.info("✓ Sankey diagram saved")

    def generate_stats(self):
        """
        Main function to generate statistics and visualizations.

        Orchestrates the full statistics generation process:
        1. Load documents from Qdrant
        2. Calculate statistics by organization
        3. Create Excel workbook with Stats sheet
        4. Create Sankey diagram visualization
        5. Embed Sankey in Stats Viz sheet
        6. Save workbook

        Returns:
            bool: True if successful, False otherwise
        """
        logger.info("=" * 60)
        logger.info("Generating Organization Statistics from Qdrant")
        logger.info("=" * 60)

        # Load documents from Qdrant
        documents = self.load_data_from_qdrant()
        if not documents:
            logger.warning("No documents found in Qdrant")
            return False

        # Calculate stats
        agency_data = self.calculate_stats(documents)

        # Create new Excel workbook
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create stats sheet
        self.create_stats_sheet(wb, agency_data)

        # Create Sankey diagram
        try:
            self.create_sankey_diagram(documents)
        except Exception as e:
            logger.error("Failed to create Sankey diagram: %s", e)
            logger.error(
                "Make sure you have plotly and kaleido installed: pip install plotly kaleido"
            )

        # Embed Sankey diagram in Stats Viz sheet
        self.create_stats_viz_sheet(wb)

        # Save workbook
        logger.info("Saving workbook...")
        wb.save(self.output_path)
        logger.info("✓ Saved: %s", self.output_path)
        logger.info("✓ Created Stats tab with organization statistics")
        logger.info("✓ Created Stats Viz tab with embedded Sankey diagram")

        logger.info("=" * 60)
        logger.info("Statistics generation complete!")
        logger.info("=" * 60)

        return True


def main():
    """
    Command-line entry point for statistics generation.

    Parses command-line arguments and executes the statistics generation
    process using data from Qdrant database.
    """
    parser = argparse.ArgumentParser(
        description="Generate organization statistics from Qdrant database"
    )
    parser.add_argument(
        "--output",
        default="./data/pipeline_stats.xlsx",
        help="Path to output Excel file (default: ./data/pipeline_stats.xlsx)",
    )

    args = parser.parse_args()

    generator = StatsGenerator(output_path=args.output)
    success = generator.generate_stats()

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
