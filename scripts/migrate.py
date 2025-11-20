import hashlib
import logging
import os
import sys
import uuid
from pathlib import Path

import openpyxl

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pipeline.db import db
from pipeline.utils import generate_doc_id

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

METADATA_PATH = "./data/pdf_metadata.xlsx"


def migrate_excel_to_qdrant():
    """Read Excel and upsert to Qdrant."""
    if not os.path.exists(METADATA_PATH):
        logger.error(f"Metadata file not found: {METADATA_PATH}")
        return

    logger.info(f"Loading metadata from: {METADATA_PATH}")
    wb = openpyxl.load_workbook(METADATA_PATH)

    logger.info(f"Available sheets: {wb.sheetnames}")

    # Try to find the best sheet
    if "PDF Metadata" in wb.sheetnames:
        ws = wb["PDF Metadata"]
    elif "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
    else:
        ws = wb.active

    logger.info(f"Using sheet: {ws.title}")

    headers = [cell.value for cell in ws[1]]
    logger.info(f"Headers: {headers}")

    # Map headers to lowercase for easier access
    # Handle None headers
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            header_map[str(h).lower().strip()] = i

    logger.info(f"Header Map: {header_map}")

    count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # Extract fields - try multiple variations
        url = None
        for key in ["url", "link", "source"]:
            if key in header_map:
                url = row[header_map[key]]
                break

        filepath = None
        for key in ["filepath", "file path", "filename", "file"]:
            if key in header_map:
                filepath = row[header_map[key]]
                break

        agency = None
        for key in ["agency", "organization", "source"]:
            if key in header_map:
                agency = row[header_map[key]]
                break

        year = None
        for key in ["year", "date", "published"]:
            if key in header_map:
                year = row[header_map[key]]
                break

        title = None
        for key in ["title", "document title", "name"]:
            if key in header_map:
                title = row[header_map[key]]
                break

        # Status fields
        download_error = None
        for key in ["download error", "error"]:
            if key in header_map:
                download_error = row[header_map[key]]
                break

        parsed_folder = None
        for key in ["parsed folder", "parsed_folder", "output folder"]:
            if key in header_map:
                parsed_folder = row[header_map[key]]
                break

        # Fallback: if no URL/Filepath but we have Title, use Title as ID source (risky but better than nothing)
        if not url and not filepath and not title:
            # logger.warning(f"Row {row_idx}: No identifier (URL/Filepath/Title), skipping.")
            continue

        # Determine status
        status = "pending"
        if download_error:
            status = "download_failed"
        elif parsed_folder:
            status = "parsed"
        elif filepath and str(filepath) != "Error Downloading":
            status = "downloaded"

        # Generate ID
        unique_key = url if url else (filepath if filepath else title)
        if not unique_key:
            continue

        doc_id = generate_doc_id(str(unique_key))

        # Extract additional metadata fields for faceting
        country = None
        for key in ["country"]:
            if key in header_map:
                country = row[header_map[key]]
                break

        region = None
        for key in ["region", "geographic scope"]:
            if key in header_map:
                region = row[header_map[key]]
                break

        evaluation_type = None
        for key in ["evaluation type", "type"]:
            if key in header_map:
                evaluation_type = row[header_map[key]]
                break

        theme = None
        for key in ["theme", "themes"]:
            if key in header_map:
                theme = row[header_map[key]]
                break

        # Construct Metadata
        metadata = {
            "url": url,
            "filepath": filepath,
            "organization": agency,  # Map 'agency' to 'organization' for consistency
            "year": year,
            "country": country,
            "region": region,
            "evaluation_type": evaluation_type,
            "theme": theme,
            "title": title,
            "status": status,
            "parsed_folder": parsed_folder,
            "download_error": download_error,
        }

        # Upsert
        db.upsert_document(doc_id, metadata)
        count += 1
        if count % 10 == 0:
            logger.info(f"Migrated {count} documents...")

    logger.info(f"Migration complete. Total documents: {count}")


if __name__ == "__main__":
    migrate_excel_to_qdrant()
