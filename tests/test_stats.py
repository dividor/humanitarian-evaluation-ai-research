"""
Test suite for stats.py - Statistics generation functionality

Tests the actual statistics pipeline components without over-mocking.
"""

import sys
import tempfile
from pathlib import Path

# Add pipeline directory to path before importing
sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

import openpyxl  # noqa: E402
from stats import StatsGenerator  # noqa: E402


class TestStatsGenerator:
    """Test StatsGenerator class functionality"""

    def test_stats_generator_initialization(self):
        """Test that stats generator initializes correctly"""
        generator = StatsGenerator()

        assert generator.metadata_path == "./data/pdf_metadata.xlsx"

    def test_stats_generator_custom_path(self):
        """Test stats generator with custom metadata path"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_path = f"{tmpdir}/test.xlsx"
            generator = StatsGenerator(metadata_path=test_path)

            assert generator.metadata_path == test_path

    def test_has_executive_summary_valid(self):
        """Test has_executive_summary with valid content"""
        generator = StatsGenerator()

        # Valid summary content
        assert (
            generator.has_executive_summary("Executive Summary section content") is True
        )
        assert (
            generator.has_executive_summary("[H1] Executive Summary\n[H2] Overview")
            is True
        )

    def test_has_executive_summary_invalid(self):
        """Test has_executive_summary with invalid content"""
        generator = StatsGenerator()

        # Invalid or empty content
        assert generator.has_executive_summary("No") is False
        assert generator.has_executive_summary("No TOC") is False
        assert generator.has_executive_summary("") is False
        assert generator.has_executive_summary(None) is False
        assert generator.has_executive_summary(123) is False

    def test_load_metadata_missing_file(self):
        """Test load_metadata with non-existent file"""
        generator = StatsGenerator(metadata_path="/nonexistent/file.xlsx")

        result = generator.load_metadata()

        assert result is None

    def test_load_metadata_valid_file(self):
        """Test loading valid metadata file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Metadata"

            # Add headers
            ws.append(
                [
                    "Agency",
                    "Year",
                    "Download Error",
                    "Parsed Folder",
                    "TOC",
                    "Key Content Sections",
                    "Abstractive Summary (map reduced)",
                    "Page Count",
                    "Word Count",
                ]
            )

            # Add sample data
            ws.append(
                [
                    "TestAgency",
                    "2024",
                    "",
                    "./parsed/test",
                    "[H1] Introduction",
                    "Executive Summary content",
                    "Summary text",
                    100,
                    5000,
                ]
            )
            wb.save(test_file)

            generator = StatsGenerator(metadata_path=str(test_file))
            result = generator.load_metadata()

            assert result is not None
            wb_loaded, ws_loaded, cols = result

            assert cols["agency"] is not None
            assert cols["year"] is not None
            assert cols["toc"] is not None

            wb_loaded.close()

    def test_load_metadata_organization_column(self):
        """Test that load_metadata handles 'Organization' column as alias for 'Agency'"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            # Use "Organization" instead of "Agency"
            ws.append(["Organization", "Year"])
            ws.append(["TestOrg", "2024"])
            wb.save(test_file)

            generator = StatsGenerator(metadata_path=str(test_file))
            result = generator.load_metadata()

            assert result is not None
            wb_loaded, ws_loaded, cols = result

            # Should map Organization to agency
            assert cols["agency"] is not None

            wb_loaded.close()

    def test_calculate_stats_single_org(self):
        """Test calculating statistics for single organization"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            ws.append(
                [
                    "Agency",
                    "Year",
                    "Download Error",
                    "Parsed Folder",
                    "TOC",
                    "Key Content Sections",
                    "Abstractive Summary (map reduced)",
                    "Page Count",
                    "Word Count",
                ]
            )

            # Add 3 documents for TestAgency
            ws.append(
                [
                    "TestAgency",
                    "2024",
                    "",
                    "./parsed/1",
                    "[H1] TOC",
                    "Summary",
                    "Abstract",
                    50,
                    2000,
                ]
            )
            ws.append(
                [
                    "TestAgency",
                    "2024",
                    "",
                    "./parsed/2",
                    "[H1] TOC",
                    "Summary",
                    "Abstract",
                    60,
                    2500,
                ]
            )
            ws.append(["TestAgency", "2023", "", "./parsed/3", "", "", "", 40, 1500])

            wb.save(test_file)

            generator = StatsGenerator(metadata_path=str(test_file))
            wb, ws, cols = generator.load_metadata()

            stats = generator.calculate_stats(ws, cols)

            assert "TestAgency" in stats
            org_stats = stats["TestAgency"]

            assert org_stats["total"] == 3
            assert org_stats["downloaded"] == 3  # No errors
            assert org_stats["parsed"] == 3  # All have parsed folder
            assert org_stats["has_toc"] == 2  # 2 have TOC
            assert org_stats["has_exec_summary"] == 2  # 2 have summaries
            assert org_stats["has_abstractive_summary"] == 2
            assert len(set(org_stats["years"])) == 2  # 2023 and 2024 (unique)
            assert org_stats["total_pages"] == 150  # 50+60+40
            assert org_stats["page_count"] == 3
            assert org_stats["total_words"] == 6000  # 2000+2500+1500
            assert org_stats["word_count"] == 3

            wb.close()

    def test_calculate_stats_multiple_orgs(self):
        """Test calculating statistics for multiple organizations"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            ws.append(["Agency", "Year", "Parsed Folder"])

            # Add documents for different agencies
            ws.append(["OrgA", "2024", "./parsed/1"])
            ws.append(["OrgA", "2024", "./parsed/2"])
            ws.append(["OrgB", "2024", "./parsed/3"])
            ws.append(["OrgB", "2023", "./parsed/4"])
            ws.append(["OrgB", "2023", ""])  # Not parsed

            wb.save(test_file)

            generator = StatsGenerator(metadata_path=str(test_file))
            wb, ws, cols = generator.load_metadata()

            stats = generator.calculate_stats(ws, cols)

            assert len(stats) == 2
            assert "OrgA" in stats
            assert "OrgB" in stats

            assert stats["OrgA"]["total"] == 2
            assert stats["OrgA"]["parsed"] == 2

            assert stats["OrgB"]["total"] == 3
            assert stats["OrgB"]["parsed"] == 2  # One not parsed

            wb.close()

    def test_calculate_stats_download_errors(self):
        """Test that download errors are counted correctly"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            ws.append(["Agency", "Year", "Download Error"])

            # Mix of successful and failed downloads
            ws.append(["TestAgency", "2024", ""])
            ws.append(["TestAgency", "2024", ""])
            ws.append(["TestAgency", "2024", "404 Error"])
            ws.append(["TestAgency", "2024", "Timeout"])

            wb.save(test_file)

            generator = StatsGenerator(metadata_path=str(test_file))
            wb, ws, cols = generator.load_metadata()

            stats = generator.calculate_stats(ws, cols)

            assert stats["TestAgency"]["total"] == 4
            assert stats["TestAgency"]["downloaded"] == 2  # Only 2 without errors

            wb.close()

    def test_create_stats_sheet(self):
        """Test creating stats sheet"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            # Create minimal metadata
            ws.append(["Agency", "Year"])
            ws.append(["TestAgency", "2024"])
            wb.save(test_file)

            # Load and calculate stats
            generator = StatsGenerator(metadata_path=str(test_file))
            wb, ws, cols = generator.load_metadata()

            agency_data = {
                "TestAgency": {
                    "years": [2024],
                    "total": 10,
                    "downloaded": 9,
                    "parsed": 8,
                    "has_toc": 7,
                    "has_exec_summary": 6,
                    "has_abstractive_summary": 5,
                    "total_pages": 500,
                    "total_words": 25000,
                    "page_count": 8,
                    "word_count": 8,
                }
            }

            generator.create_stats_sheet(wb, agency_data)

            # Check that Stats sheet was created
            assert "Stats" in wb.sheetnames

            stats_ws = wb["Stats"]

            # Check headers
            headers = [cell.value for cell in stats_ws[1]]
            assert "Organization" in headers
            assert "Total Number of PDFs" in headers
            assert "% PDFs Parsed Successfully" in headers

            # Check data row
            assert stats_ws.cell(2, 1).value == "TestAgency"
            assert stats_ws.cell(2, 4).value == 10  # Total PDFs

            wb.close()

    def test_create_stats_sheet_replaces_existing(self):
        """Test that create_stats_sheet replaces existing Stats sheet"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Agency"])

            # Create an existing Stats sheet
            wb.create_sheet("Stats")
            old_stats = wb["Stats"]
            old_stats.append(["Old Data"])

            wb.save(test_file)

            # Load and recreate stats
            wb = openpyxl.load_workbook(test_file)

            agency_data = {
                "NewAgency": {
                    "years": [2024],
                    "total": 5,
                    "downloaded": 5,
                    "parsed": 5,
                    "has_toc": 5,
                    "has_exec_summary": 5,
                    "has_abstractive_summary": 5,
                    "total_pages": 250,
                    "total_words": 12500,
                    "page_count": 5,
                    "word_count": 5,
                }
            }

            generator = StatsGenerator()
            generator.create_stats_sheet(wb, agency_data)

            # Check that new Stats sheet has new data
            stats_ws = wb["Stats"]
            data_row_value = stats_ws.cell(2, 1).value

            assert data_row_value == "NewAgency"
            assert "Old Data" not in [
                cell.value for row in stats_ws.iter_rows() for cell in row
            ]

            wb.close()

    def test_create_stats_viz_sheet_missing_image(self):
        """Test create_stats_viz_sheet with missing image file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Agency"])
            wb.save(test_file)

            wb = openpyxl.load_workbook(test_file)

            generator = StatsGenerator()
            # Should not raise error, just log warning
            generator.create_stats_viz_sheet(wb, image_path="/nonexistent/image.png")

            # Stats Viz sheet should still be created
            assert "Stats Viz" in wb.sheetnames

            wb.close()


class TestStatsConstants:
    """Test constants and configuration"""

    def test_toc_incompleteness_ratio(self):
        """Test that TOC incompleteness ratio is defined"""
        import sys
        from pathlib import Path

        sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

        from stats import TOC_INCOMPLETENESS_RATIO

        assert TOC_INCOMPLETENESS_RATIO is not None
        assert isinstance(TOC_INCOMPLETENESS_RATIO, (int, float))
        assert TOC_INCOMPLETENESS_RATIO > 0


class TestStatsIntegration:
    """Integration tests for full stats generation"""

    def test_generate_stats_end_to_end(self):
        """Test complete stats generation workflow"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            # Create complete metadata
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Metadata"

            ws.append(
                [
                    "Agency",
                    "Year",
                    "Download Error",
                    "Parsed Folder",
                    "TOC",
                    "Key Content Sections",
                    "Abstractive Summary (map reduced)",
                    "Page Count",
                    "Word Count",
                ]
            )

            # Add sample data for two organizations
            ws.append(
                [
                    "OrgA",
                    "2024",
                    "",
                    "./parsed/1",
                    "[H1] TOC",
                    "Summary",
                    "Abstract",
                    50,
                    2000,
                ]
            )
            ws.append(
                [
                    "OrgA",
                    "2023",
                    "",
                    "./parsed/2",
                    "[H1] TOC",
                    "Summary",
                    "Abstract",
                    60,
                    2500,
                ]
            )
            ws.append(
                ["OrgB", "2024", "", "./parsed/3", "[H1] TOC", "Summary", "", 40, 1500]
            )

            wb.save(test_file)

            # Generate stats (without creating Sankey diagram to avoid dependencies)
            generator = StatsGenerator(metadata_path=str(test_file))

            # Load metadata
            result = generator.load_metadata()
            assert result is not None

            wb, ws, cols = result

            # Calculate stats
            agency_data = generator.calculate_stats(ws, cols)
            assert len(agency_data) == 2

            # Create stats sheet
            generator.create_stats_sheet(wb, agency_data)
            assert "Stats" in wb.sheetnames

            # Verify stats sheet content
            stats_ws = wb["Stats"]
            row_count = stats_ws.max_row

            # Should have header + 2 data rows
            assert row_count == 3

            wb.close()


def test_module_imports():
    """Test that all required modules can be imported"""
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

    from stats import StatsGenerator

    assert StatsGenerator is not None


def test_stats_sheet_column_headers():
    """Test that expected column headers are defined correctly"""
    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Agency"])
        wb.save(test_file)

        wb = openpyxl.load_workbook(test_file)

        agency_data = {
            "Test": {
                "years": [2024],
                "total": 1,
                "downloaded": 1,
                "parsed": 1,
                "has_toc": 1,
                "has_exec_summary": 1,
                "has_abstractive_summary": 1,
                "total_pages": 50,
                "total_words": 2000,
                "page_count": 1,
                "word_count": 1,
            }
        }

        generator = StatsGenerator()
        generator.create_stats_sheet(wb, agency_data)

        stats_ws = wb["Stats"]
        headers = [cell.value for cell in stats_ws[1]]

        # Verify all expected headers are present
        expected_headers = [
            "Organization",
            "Start Year",
            "End Year",
            "Total Number of PDFs",
            "% PDFs Downloaded Successfully",
            "Number of PDFs Downloaded Successfully",
            "% PDFs Parsed Successfully",
            "Number of PDFs Parsed Successfully",
            "% PDFs with Table of Contents",
            "Number of PDFs with Table of Contents",
            "% PDFs with Summary Section",
            "Number of PDFs with Summary Section",
            "% PDFs with Abstractive Summary",
            "Number of PDFs with Abstractive Summary",
            "Average Number of Pages per PDF",
            "Average Number of Words per PDF",
        ]

        for expected in expected_headers:
            assert expected in headers, f"Missing header: {expected}"

        wb.close()


# noqa: E402
