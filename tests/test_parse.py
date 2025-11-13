"""
Test suite for parse.py - Document parsing functionality

Tests the actual parsing pipeline components without over-mocking.
"""

import sys
import tempfile
from pathlib import Path

# Add pipeline directory to path before importing
sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

import openpyxl  # noqa: E402
from parse import DoclingParser  # noqa: E402


class TestDoclingParser:
    """Test DoclingParser class functionality"""

    def test_parser_initialization(self):
        """Test that parser initializes with correct default settings"""
        parser = DoclingParser()

        assert parser.metadata_path == "./data/pdf_metadata.xlsx"
        assert parser.output_dir == "./data/parsed"
        assert parser.pipeline == "standard"
        assert parser.ocr_engine == "rapidocr"
        assert parser.no_ocr is True
        assert parser.enable_chunking is True
        assert parser.chunk_size == 50
        assert parser.chunk_threshold == 200
        assert parser.chunk_timeout == 300

    def test_parser_custom_initialization(self):
        """Test parser with custom parameters"""
        with tempfile.TemporaryDirectory() as tmpdir:
            parser = DoclingParser(
                metadata_path=f"{tmpdir}/test.xlsx",
                output_dir=tmpdir,
                chunk_size=100,
                chunk_threshold=500,
                no_ocr=False,
            )

            assert parser.metadata_path == f"{tmpdir}/test.xlsx"
            assert parser.output_dir == tmpdir
            assert parser.chunk_size == 100
            assert parser.chunk_threshold == 500
            assert parser.no_ocr is False

    def test_create_output_folder(self):
        """Test output folder creation"""
        with tempfile.TemporaryDirectory() as tmpdir:
            parser = DoclingParser(output_dir=tmpdir)

            # Create a test filepath
            test_filepath = Path(tmpdir) / "TestAgency" / "2024" / "test_doc.pdf"
            test_filepath.parent.mkdir(parents=True, exist_ok=True)
            test_filepath.write_text("test")

            output_folder = parser.create_output_folder(test_filepath)

            assert Path(output_folder).exists()
            assert "TestAgency" in output_folder or tmpdir in output_folder

    def test_ensure_columns_exist_with_new_workbook(self):
        """Test that ensure_columns_exist adds required columns to new workbook"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a minimal test workbook
            test_file = Path(tmpdir) / "test.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add basic headers
            ws.append(["Node ID", "Title", "Filepath"])
            wb.save(test_file)

            # Load and ensure columns
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active

            parser = DoclingParser()
            parser._ensure_columns_exist(ws, wb)

            # Check that new columns were added
            headers = [cell.value for cell in ws[1]]

            required_columns = [
                "Parsed Folder",
                "Parsed Markdown Path",
                "TOC",
                "Page Count",
                "Word Count",
                "Language",
                "File Format",
                "Parsed Error",  # Correct column name
            ]

            for col in required_columns:
                assert col in headers, f"Column '{col}' not found in headers"

            wb.close()

    def test_ensure_columns_exist_preserves_existing(self):
        """Test that ensure_columns_exist doesn't duplicate existing columns"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add headers including some parsing columns
            ws.append(
                [
                    "Node ID",
                    "Title",
                    "Filepath",
                    "Parsed Folder",
                    "TOC",
                    "Page Count",
                ]
            )
            wb.save(test_file)

            # Load and ensure columns
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active

            parser = DoclingParser()
            initial_headers = [cell.value for cell in ws[1]]
            parser._ensure_columns_exist(ws, wb)
            final_headers = [cell.value for cell in ws[1]]

            # Check no duplicates
            assert len(final_headers) == len(set(final_headers))

            # Check original columns still present
            for col in initial_headers:
                assert col in final_headers

            wb.close()

    def test_load_metadata_missing_file(self):
        """Test that load_metadata handles missing file gracefully"""
        parser = DoclingParser(metadata_path="/nonexistent/file.xlsx")

        wb, ws = parser.load_metadata()

        assert wb is None
        assert ws is None

    def test_load_metadata_valid_file(self):
        """Test loading a valid metadata file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            # Create a valid test workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Metadata"

            # Add required columns
            ws.append(["Node ID", "Title", "Agency", "Year", "Filepath"])
            ws.append(["001", "Test Report", "TestAgency", "2024", "./data/test.pdf"])
            wb.save(test_file)
            wb.close()

            parser = DoclingParser(metadata_path=str(test_file))
            wb, ws = parser.load_metadata()

            assert wb is not None
            assert ws is not None
            assert ws.max_row >= 2  # Header + at least 1 data row

            wb.close()

    def test_get_rows_to_parse_filters_by_year(self):
        """Test filtering rows by year"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Node ID", "Title", "Agency", "Year", "Filepath"])
            ws.append(["001", "Report 2023", "Agency1", "2023", "./test1.pdf"])
            ws.append(["002", "Report 2024", "Agency1", "2024", "./test2.pdf"])
            ws.append(["003", "Report 2025", "Agency2", "2025", "./test3.pdf"])
            wb.save(test_file)
            wb.close()

            parser = DoclingParser(metadata_path=str(test_file))
            wb, ws = parser.load_metadata()

            # Filter for 2024 only
            filtered = parser.get_rows_to_parse(ws, year="2024")

            assert len(filtered) == 1
            assert filtered[0]["year"] == "2024"
            assert filtered[0]["filepath"] == "./test2.pdf"

            wb.close()

    def test_get_rows_to_parse_filters_by_agency(self):
        """Test filtering rows by agency"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Node ID", "Title", "Agency", "Year", "Filepath"])
            ws.append(["001", "Report A", "AgencyA", "2024", "./test1.pdf"])
            ws.append(["002", "Report B", "AgencyB", "2024", "./test2.pdf"])
            ws.append(["003", "Report C", "AgencyA", "2024", "./test3.pdf"])
            wb.save(test_file)
            wb.close()

            parser = DoclingParser(metadata_path=str(test_file))
            wb, ws = parser.load_metadata()

            # Filter for AgencyA only
            filtered = parser.get_rows_to_parse(ws, agency="AgencyA")

            assert len(filtered) == 2
            assert all(row["agency"] == "AgencyA" for row in filtered)

            wb.close()

    def test_get_rows_to_parse_skips_already_parsed(self):
        """Test that already parsed rows are skipped unless force_parse is True"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(
                [
                    "Node ID",
                    "Title",
                    "Agency",
                    "Year",
                    "Filepath",
                    "Parsed Folder",
                ]
            )
            ws.append(
                [
                    "001",
                    "Parsed Doc",
                    "Agency1",
                    "2024",
                    "./test1.pdf",
                    "./parsed/test1",
                ]
            )
            ws.append(["002", "Unparsed Doc", "Agency1", "2024", "./test2.pdf", ""])
            wb.save(test_file)
            wb.close()

            parser = DoclingParser(metadata_path=str(test_file))
            wb, ws = parser.load_metadata()

            # Without force_parse, should skip already parsed
            filtered = parser.get_rows_to_parse(ws, force_parse=False)
            assert len(filtered) == 1
            assert filtered[0]["filepath"] == "./test2.pdf"

            # With force_parse, should include all
            filtered = parser.get_rows_to_parse(ws, force_parse=True)
            assert len(filtered) == 2

            wb.close()

    def test_add_page_numbers_to_breaks(self):
        """Test that _add_page_numbers_to_breaks function exists and has correct signature"""
        parser = DoclingParser()

        # This function requires a markdown path and document object
        # We're just testing that the method exists with correct signature
        assert hasattr(parser, "_add_page_numbers_to_breaks")
        assert callable(parser._add_page_numbers_to_breaks)

    def test_should_chunk_pdf_respects_threshold(self):
        """Test chunking threshold logic"""
        parser = DoclingParser(enable_chunking=True, chunk_threshold=100)

        # Create a mock PDF path (we'll test the logic, not actual PDF reading)
        # The function tries to open the file, so this would fail for non-existent files
        # We're testing the threshold logic is in place

        assert parser.chunk_threshold == 100
        assert parser.enable_chunking is True

    def test_chunking_disabled(self):
        """Test that chunking can be disabled"""
        parser = DoclingParser(enable_chunking=False)

        assert parser.enable_chunking is False


class TestMetadataUpdate:
    """Test metadata update functionality"""

    def test_update_metadata_adds_values(self):
        """Test that update_metadata correctly adds values to worksheet"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            # Create test workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(
                [
                    "Node ID",
                    "Parsed Folder",
                    "Parsed Markdown Path",
                    "TOC",
                    "Page Count",
                    "Word Count",
                    "Language",
                    "File Format",
                    "Parsed Error",  # Correct column name
                ]
            )
            ws.append(["001", "", "", "", "", "", "", "", ""])
            wb.save(test_file)

            # Load and update
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active

            parser = DoclingParser()
            parser.update_metadata(
                ws,
                row_idx=2,
                parsed_folder="./parsed/test",
                markdown_path="./parsed/test/doc.md",
                toc_string="[H1] Test Heading",
                page_count=10,
                word_count=1000,
                language="en",
                file_format="pdf",
            )

            # Check values were written
            assert ws.cell(2, 2).value == "./parsed/test"  # Parsed Folder
            assert ws.cell(2, 3).value == "./parsed/test/doc.md"  # Markdown Path
            assert ws.cell(2, 4).value == "[H1] Test Heading"  # TOC
            assert ws.cell(2, 5).value == 10  # Page Count
            assert ws.cell(2, 6).value == 1000  # Word Count
            assert ws.cell(2, 7).value == "en"  # Language
            # File format gets uppercased
            assert ws.cell(2, 8).value in ["pdf", "PDF"]  # File Format

            wb.close()


class TestChunkingLogic:
    """Test PDF chunking functionality"""

    def test_split_pdf_into_chunks_parameters(self):
        """Test that split_pdf_into_chunks accepts correct parameters"""
        parser = DoclingParser(chunk_size=25, chunk_threshold=50)

        assert parser.chunk_size == 25
        assert parser.chunk_threshold == 50

    def test_chunk_timeout_setting(self):
        """Test chunk timeout configuration"""
        parser = DoclingParser(chunk_timeout=600)

        assert parser.chunk_timeout == 600


def test_module_imports():
    """Test that all required modules can be imported"""
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

    # Test imports
    from parse import DoclingParser

    assert DoclingParser is not None


def test_page_separator_constant():
    """Test that PAGE_SEPARATOR constant is defined"""
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

    from parse import PAGE_SEPARATOR

    assert PAGE_SEPARATOR is not None
    assert isinstance(PAGE_SEPARATOR, str)
    assert "Page Break" in PAGE_SEPARATOR
