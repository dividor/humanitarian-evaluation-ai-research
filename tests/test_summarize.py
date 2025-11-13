"""
Test suite for summarize.py - Document summarization functionality

Tests the actual summarization pipeline components without over-mocking.
"""

import tempfile
from pathlib import Path

import openpyxl

from pipeline.summarize import (
    DocumentSummarizer,
    clean_markdown_formatting,
    get_model_type,
)


class TestDocumentSummarizer:
    """Test DocumentSummarizer class functionality"""

    def test_summarizer_initialization(self):
        """Test that summarizer initializes correctly"""
        summarizer = DocumentSummarizer()

        assert summarizer.metadata_path == "./data/pdf_metadata.xlsx"
        assert summarizer.model is None  # Not loaded until needed
        assert summarizer.hf_token is None  # Not loaded until needed

    def test_summarizer_custom_path(self):
        """Test summarizer with custom metadata path"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_path = f"{tmpdir}/test.xlsx"
            summarizer = DocumentSummarizer(metadata_path=test_path)

            assert summarizer.metadata_path == test_path

    def test_ensure_columns_exist(self):
        """Test that ensure_columns_exist adds required summary columns"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            # Create minimal workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Node ID", "Title"])
            wb.save(test_file)

            # Load and ensure columns
            wb = openpyxl.load_workbook(test_file)
            ws = wb.active

            summarizer = DocumentSummarizer()
            summarizer.ensure_columns_exist(ws)

            headers = [cell.value for cell in ws[1]]

            required_columns = [
                "Key Content Sections",
                "Centroid Summary",
                "Abstractive Summary (map reduced)",
                "Abstractive Summary Input Method",
            ]

            for col in required_columns:
                assert col in headers, f"Column '{col}' not found"

            wb.close()

    def test_ensure_columns_preserves_existing(self):
        """Test that ensure_columns_exist doesn't duplicate columns"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(
                [
                    "Node ID",
                    "Title",
                    "Key Content Sections",
                    "Centroid Summary",
                ]
            )
            wb.save(test_file)

            wb = openpyxl.load_workbook(test_file)
            ws = wb.active

            summarizer = DocumentSummarizer()
            summarizer.ensure_columns_exist(ws)
            final_headers = [cell.value for cell in ws[1]]

            # No duplicates
            assert len(final_headers) == len(set(final_headers))

            wb.close()

    def test_load_metadata_missing_file(self):
        """Test load_metadata with non-existent file"""
        summarizer = DocumentSummarizer(metadata_path="/nonexistent/file.xlsx")

        result = summarizer.load_metadata()

        assert result is None

    def test_load_metadata_valid_file(self):
        """Test loading valid metadata file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = Path(tmpdir) / "test.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Metadata"
            ws.append(
                [
                    "Node ID",
                    "Title",
                    "Parsed Markdown Path",
                    "TOC",
                    "Language",
                ]
            )
            ws.append(
                [
                    "001",
                    "Test Doc",
                    "./parsed/test/doc.md",
                    "[H1] Introduction",
                    "en",
                ]
            )
            wb.save(test_file)

            summarizer = DocumentSummarizer(metadata_path=str(test_file))
            result = summarizer.load_metadata()

            assert result is not None
            wb, ws, rows = result
            assert len(rows) == 1
            assert rows[0]["node_id"] == "001"
            assert rows[0]["title"] == "Test Doc"

            wb.close()

    def test_save_content_to_file(self):
        """Test saving content to file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a markdown file path
            md_path = Path(tmpdir) / "test" / "document.md"
            md_path.parent.mkdir(parents=True, exist_ok=True)
            md_path.write_text("# Test Document")

            summarizer = DocumentSummarizer()
            content = "This is test summary content."

            result_path = summarizer.save_content_to_file(
                str(md_path), content, "test_summary"
            )

            assert result_path is not None
            assert "test_summary.txt" in result_path

            # Verify file was created and contains content
            full_path = Path(tmpdir) / "test" / "document_test_summary.txt"
            assert full_path.exists()
            saved_content = full_path.read_text()
            assert "This is test summary content" in saved_content

    def test_load_content_from_file(self):
        """Test loading content from saved file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test file
            test_file = Path(tmpdir) / "test_content.txt"
            test_content = "This is test content\nwith multiple lines."
            test_file.write_text(test_content)

            summarizer = DocumentSummarizer()
            loaded_content = summarizer.load_content_from_file(str(test_file))

            assert loaded_content == test_content

    def test_load_content_from_missing_file(self):
        """Test loading content from non-existent file"""
        summarizer = DocumentSummarizer()
        result = summarizer.load_content_from_file("/nonexistent/file.txt")

        assert result is None

    def test_detect_summary_sections_with_toc(self):
        """Test detecting summary sections from TOC"""
        summarizer = DocumentSummarizer()

        toc_with_summary = """
[H1|p.1] Executive Summary
[H1|p.5] Introduction
[H2|p.7] Background
[H1|p.10] Main Findings
[H1|p.15] Conclusions
        """

        sections = summarizer.detect_summary_sections(toc_with_summary)

        # Should detect Executive Summary and Conclusions
        assert sections is not None
        assert len(sections) > 0

    def test_detect_summary_sections_without_summary(self):
        """Test detecting summary sections from TOC without summaries"""
        summarizer = DocumentSummarizer()

        toc_no_summary = """
[H1|p.1] Introduction
[H2|p.3] Methodology
[H1|p.5] Analysis
        """

        sections = summarizer.detect_summary_sections(toc_no_summary)

        # Should return empty or None
        assert sections is None or len(sections) == 0

    def test_tokenize_sentences(self):
        """Test sentence tokenization"""
        summarizer = DocumentSummarizer()

        text = "This is sentence one. This is sentence two! And this is sentence three?"

        sentences = summarizer.tokenize_sentences(text)

        assert len(sentences) >= 3
        assert any("sentence one" in s for s in sentences)
        assert any("sentence two" in s for s in sentences)

    def test_tokenize_sentences_empty(self):
        """Test tokenizing empty text"""
        summarizer = DocumentSummarizer()

        sentences = summarizer.tokenize_sentences("")

        assert len(sentences) == 0


class TestUtilityFunctions:
    """Test utility functions"""

    def test_clean_markdown_formatting(self):
        """Test markdown cleaning function"""
        markdown_text = """
# Heading 1

![Image](image.png)

------- Page Break -------

Some text here.

<!-- HTML comment -->

More text.
        """

        cleaned = clean_markdown_formatting(markdown_text)

        # Images should be removed
        assert "![Image]" not in cleaned

        # Page breaks should be removed
        assert "Page Break" not in cleaned

        # HTML comments should be removed
        assert "<!--" not in cleaned
        assert "HTML comment" not in cleaned

        # Regular text should remain
        assert "Some text here" in cleaned
        assert "More text" in cleaned

    def test_clean_markdown_empty(self):
        """Test cleaning empty markdown"""
        result = clean_markdown_formatting("")

        assert result == ""

    def test_get_model_type_bart(self):
        """Test model type detection for BART"""
        bart_model = {"model": "facebook/bart-large-cnn"}

        model_type = get_model_type(bart_model)

        assert model_type == "bart"

    def test_get_model_type_mistral(self):
        """Test model type detection for Mistral"""
        mistral_model = {"model": "mistralai/Mistral-7B-Instruct-v0.2"}

        model_type = get_model_type(mistral_model)

        assert model_type == "mistral"

    def test_get_model_type_llama(self):
        """Test model type detection for Llama"""
        llama_model = {"model": "meta-llama/Llama-3.1-8B-Instruct:novita"}

        model_type = get_model_type(llama_model)

        assert model_type == "llama"

    def test_get_model_type_chat(self):
        """Test model type detection for chat models"""
        chat_model = {"model": "other/chat-model"}

        model_type = get_model_type(chat_model)

        assert model_type == "chat"


class TestMarkdownLoading:
    """Test markdown loading functionality"""

    def test_load_markdown_valid_file(self):
        """Test loading valid markdown file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            md_file = Path(tmpdir) / "test.md"
            test_content = "# Test Document\n\nThis is test content."
            md_file.write_text(test_content)

            summarizer = DocumentSummarizer()
            content = summarizer.load_markdown(str(md_file))

            assert content is not None
            assert "Test Document" in content
            assert "test content" in content

    def test_load_markdown_missing_file(self):
        """Test loading non-existent markdown file"""
        summarizer = DocumentSummarizer()

        result = summarizer.load_markdown("/nonexistent/file.md")

        assert result is None


class TestExtractSummarySection:
    """Test summary section extraction"""

    def test_extract_summary_section_with_valid_toc(self):
        """Test extracting summary section with valid TOC and content"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create markdown file
            md_file = Path(tmpdir) / "test.md"
            content = """
# Executive Summary

This is the executive summary content.
It has multiple sentences.

# Introduction

This is the introduction.
            """
            md_file.write_text(content)

            # Create TOC
            toc = "[H1|p.1] Executive Summary\n[H1|p.3] Introduction"

            summarizer = DocumentSummarizer()
            result = summarizer.extract_summary_section(
                str(md_file), toc, ["Executive Summary"]
            )

            assert result is not None
            assert "executive summary content" in result.lower()

    def test_extract_summary_section_no_sections(self):
        """Test extracting when no summary sections exist"""
        with tempfile.TemporaryDirectory() as tmpdir:
            md_file = Path(tmpdir) / "test.md"
            content = "# Introduction\n\nThis is the introduction."
            md_file.write_text(content)

            toc = "[H1|p.1] Introduction"

            summarizer = DocumentSummarizer()
            result = summarizer.extract_summary_section(str(md_file), toc, [])

            assert result is None


def test_module_constants():
    """Test that required constants are defined"""
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

    from summarize import (
        LLM_MODEL,
        LLM_REDUCTION_PROMPT,
        LLM_SUMMARY_PROMPT,
        NUM_CENTROID_SENTENCES,
    )

    assert LLM_MODEL is not None
    assert isinstance(LLM_MODEL, dict)
    assert "model" in LLM_MODEL

    assert NUM_CENTROID_SENTENCES > 0
    assert isinstance(LLM_REDUCTION_PROMPT, str)
    assert isinstance(LLM_SUMMARY_PROMPT, str)


def test_module_imports():
    """Test that all required modules can be imported"""
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent / "pipeline"))

    from summarize import DocumentSummarizer  # noqa: E402

    assert DocumentSummarizer is not None


# noqa: E402
