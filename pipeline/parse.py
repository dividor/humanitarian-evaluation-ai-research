"""
parse.py - Parse PDFs using Docling Python API, with metadata tracking
"""

import argparse
import logging
import multiprocessing
import os
import signal
import sys

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import tempfile
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from docling.backend.docling_parse_v2_backend import DoclingParseV2DocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import (
    PdfPipelineOptions,
    TableFormerMode,
    TableStructureOptions,
)
from docling.document_converter import DocumentConverter, PdfFormatOption
from docling.pipeline.standard_pdf_pipeline import StandardPdfPipeline
from docling_core.types.doc.base import ImageRefMode
from hierarchical.postprocessor import ResultPostprocessor

from pipeline.db import db
from pipeline.utils import generate_doc_id

# Page separator pattern for markdown output
# Note: Docling's page_break_placeholder doesn't support dynamic page numbers yet
# This is a static separator that appears between pages
PAGE_SEPARATOR = "\n\n------- Page Break -------\n\n"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


def _parse_pdf_worker(filepath, output_folder, parser_args, result_queue):
    """
    Worker function to parse PDF in a separate process.

    Used for subprocess-based parsing with OOM protection. Recreates the parser
    in an isolated process and returns results via a queue.

    Args:
        filepath: Path to the PDF file to parse
        output_folder: Destination folder for parsed output
        parser_args: Dictionary of arguments to recreate DoclingParser
        result_queue: Multiprocessing queue for returning results
    """
    try:
        # Recreate parser in this process
        parser = DoclingParser(**parser_args)
        result = parser.parse_pdf_with_docling(filepath, output_folder)
        result_queue.put({"success": True, "result": result})
    except Exception as e:
        import traceback

        result_queue.put(
            {"success": False, "error": str(e), "traceback": traceback.format_exc()}
        )


class DoclingParser:
    """
    PDF/DOCX document parser using Docling with metadata tracking.

    This class handles parsing of documents from a metadata Excel file, extracting
    structured content, table of contents, and language information. Supports
    automatic chunking for large PDFs and OOM protection through subprocess isolation.
    """

    def __init__(
        self,
        metadata_path="./data/pdf_metadata.xlsx",
        output_dir="./data/parsed",
        pipeline="standard",
        ocr_engine="rapidocr",
        ocr_lang="en,fr,es,pt",
        pdf_backend="dlparse_v4",
        table_mode="fast",
        no_ocr=True,
        enrich_picture_description=False,
        images_scale=1.0,
        enable_chunking=True,
        chunk_size=50,
        chunk_threshold=200,
        chunk_timeout=300,
    ):
        """
        Initialize the DoclingParser with configuration options.

        Args:
            metadata_path: Path to Excel file containing document metadata
            output_dir: Directory where parsed documents will be saved
            pipeline: Docling pipeline type (default: 'standard')
            ocr_engine: OCR engine to use ('rapidocr' or 'easyocr')
            ocr_lang: Comma-separated list of OCR languages (e.g., 'en,fr,es,pt')
            pdf_backend: PDF backend version (default: 'dlparse_v4')
            table_mode: Table extraction mode ('fast' or 'accurate')
            no_ocr: If True, disables OCR for faster processing
            enrich_picture_description: If True, enables AI picture descriptions
            images_scale: Image resolution scale (1.0 = normal, 2.0-3.0 = higher quality)
            enable_chunking: If True, enables automatic chunking for large PDFs
            chunk_size: Number of pages per chunk when chunking is enabled
            chunk_threshold: Minimum page count to trigger chunking
            chunk_timeout: Timeout in seconds for processing each chunk
        """
        self.metadata_path = metadata_path
        self.output_dir = output_dir
        self.pipeline = pipeline
        self.ocr_engine = ocr_engine
        self.ocr_lang = ocr_lang
        self.pdf_backend = pdf_backend
        self.table_mode = table_mode
        self.no_ocr = no_ocr
        self.enrich_picture_description = enrich_picture_description
        self.images_scale = images_scale
        self.enable_chunking = enable_chunking
        self.chunk_size = chunk_size
        self.chunk_threshold = chunk_threshold
        self.chunk_timeout = chunk_timeout

        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

        # Initialize DocumentConverter with pipeline options
        pipeline_options = PdfPipelineOptions(
            do_ocr=not no_ocr,
            do_table_structure=True,
            table_structure_options=TableStructureOptions(
                mode=(
                    TableFormerMode.FAST
                    if table_mode == "fast"
                    else TableFormerMode.ACCURATE
                )
            ),
            images_scale=images_scale,
            generate_page_images=False,
            generate_picture_images=True,
        )

        logger.info(
            "Initializing Docling converter (this may take a few minutes on first run)..."
        )
        self.converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_cls=StandardPdfPipeline,
                    pipeline_options=pipeline_options,
                    backend=DoclingParseV2DocumentBackend,
                ),
                # DOCX support (uses default converter, no special options needed)
                InputFormat.DOCX: None,
            }
        )
        logger.info("✓ Docling converter initialized")

    def load_metadata(self):
        """
        Load metadata from Excel file.

        Returns:
            tuple: (workbook, worksheet) if successful, (None, None) if file not found
        """
        if not os.path.exists(self.metadata_path):
            logger.error(f"Metadata file not found: {self.metadata_path}")
            return None, None

        logger.info(f"Loading metadata from: {self.metadata_path}")
        wb = openpyxl.load_workbook(self.metadata_path)
        ws = wb.active
        return wb, ws

    def get_rows_to_parse(self, ws, year=None, agency=None, force_parse=False):
        """
        Get rows that need parsing based on filters and existing parsed status.

        Args:
            ws: Excel worksheet containing metadata
            year: Optional year filter to limit parsing to specific year
            agency: Optional agency filter to limit parsing to specific agency
            force_parse: If True, re-parse documents even if already parsed

        Returns:
            list: List of dictionaries containing row data (row_idx, filepath, year, agency)
        """
        headers = [cell.value for cell in ws[1]]

        # Find column indices - use title case names, with fallback to old names
        filepath_col = None
        if "Filepath" in headers:
            filepath_col = headers.index("Filepath") + 1
        elif "filepath" in headers:
            filepath_col = headers.index("filepath") + 1

        if not filepath_col:
            logger.error("Required column 'Filepath' or 'filepath' not found")
            return []

        # Optional columns
        year_col = headers.index("Year") + 1 if "Year" in headers else None
        agency_col = headers.index("Agency") + 1 if "Agency" in headers else None

        # Look for Download Error column
        download_error_col = None
        if "Download Error" in headers:
            download_error_col = headers.index("Download Error") + 1
        elif "download_error" in headers:
            download_error_col = headers.index("download_error") + 1

        # Look for Parsed Folder (new) or parsed_folder (old)
        parsed_folder_col = None
        if "Parsed Folder" in headers:
            parsed_folder_col = headers.index("Parsed Folder") + 1
        elif "parsed_folder" in headers:
            parsed_folder_col = headers.index("parsed_folder") + 1

        # Look for Parsed Error (new) or parsing_error (old)
        parsed_error_col = None
        if "Parsed Error" in headers:
            parsed_error_col = headers.index("Parsed Error") + 1
        elif "parsing_error" in headers:
            parsed_error_col = headers.index("parsing_error") + 1

        rows_to_parse = []

        for row_idx in range(2, ws.max_row + 1):
            filepath = ws.cell(row_idx, filepath_col).value
            if not filepath:
                continue

            # Skip rows where download failed (filepath is "Error Downloading")
            if str(filepath).strip() == "Error Downloading":
                logger.debug(
                    f"Skipping row {row_idx} - download failed (filepath='Error Downloading')"
                )
                continue

            # Skip if there's a download error in the Download Error column
            if download_error_col:
                download_error = ws.cell(row_idx, download_error_col).value
                if download_error and str(download_error).strip():
                    logger.debug(
                        f"Skipping {filepath} - has download error: {download_error}"
                    )
                    continue

            # Check year filter
            if year and year_col:
                row_year = ws.cell(row_idx, year_col).value
                if row_year != year:
                    continue

            # Check agency filter
            if agency and agency_col:
                row_agency = ws.cell(row_idx, agency_col).value
                if row_agency != agency:
                    continue

            # Check if already parsed (unless force_parse is enabled)
            if not force_parse:
                # Skip if there's a parsing error
                if parsed_error_col:
                    error_value = ws.cell(row_idx, parsed_error_col).value
                    if error_value and str(error_value).strip():
                        logger.debug(f"Skipping {filepath} - has parsing error")
                        continue

                # Skip if already parsed successfully
                if parsed_folder_col:
                    parsed_value = ws.cell(row_idx, parsed_folder_col).value
                    if parsed_value and str(parsed_value).strip():
                        logger.debug(f"Skipping {filepath} - already parsed")
                        continue

            rows_to_parse.append(
                {
                    "row_idx": row_idx,
                    "filepath": filepath,
                    "year": (ws.cell(row_idx, year_col).value if year_col else None),
                    "agency": (
                        ws.cell(row_idx, agency_col).value if agency_col else None
                    ),
                }
            )

        logger.info(f"Found {len(rows_to_parse)} documents to parse")
        return rows_to_parse

    def create_output_folder(self, filepath):
        """
        Create output folder matching PDF structure: agency/year/document_folder.

        Parses filepath to extract agency and year, then creates a structured
        output directory. Assumes structure: data/pdfs/Agency_Name/Year/document.pdf

        Args:
            filepath: Path to the source PDF/DOCX file

        Returns:
            str: Path to the created output folder
        """
        # Parse filepath to extract agency and year
        # Assumes structure: data/pdfs/Agency_Name/Year/document.pdf
        parts = Path(filepath).parts

        try:
            # Find 'pdfs' in path
            pdfs_idx = parts.index("pdfs")
            agency = parts[pdfs_idx + 1] if len(parts) > pdfs_idx + 1 else "unknown"
            year = parts[pdfs_idx + 2] if len(parts) > pdfs_idx + 2 else "unknown"
            filename = Path(filepath).stem

            # Create agency/year/document_folder structure
            output_folder = Path(self.output_dir) / agency / year / filename
            output_folder.mkdir(parents=True, exist_ok=True)

            logger.debug(f"Created output folder: {output_folder}")
            return str(output_folder)
        except (ValueError, IndexError):
            # Fallback if structure doesn't match
            filename = Path(filepath).stem
            output_folder = Path(self.output_dir) / "other" / filename
            output_folder.mkdir(parents=True, exist_ok=True)
            logger.warning(f"Using fallback folder structure: {output_folder}")
            return str(output_folder)

    def create_pdf_symlink(self, filepath, output_folder):
        """
        Create a symbolic link to the original PDF in the output folder.

        Args:
            filepath: Path to the source PDF/DOCX file
            output_folder: Destination folder for the symlink
        """
        try:
            pdf_name = Path(filepath).name
            symlink_path = Path(output_folder) / pdf_name

            # Remove existing symlink if it exists
            if symlink_path.is_symlink() or symlink_path.exists():
                symlink_path.unlink()

            # Create symlink with absolute path
            abs_filepath = Path(filepath).resolve()
            symlink_path.symlink_to(abs_filepath)
            logger.debug(f"Created symlink: {symlink_path} -> {abs_filepath}")
        except (OSError, IOError) as e:
            logger.warning(f"Failed to create symlink: {e}")

    def should_chunk_pdf(self, filepath):
        """
        Determine if PDF should be chunked based on page count.

        Args:
            filepath: Path to the PDF file

        Returns:
            bool: True if PDF should be chunked (page count >= threshold), False otherwise
        """
        if not self.enable_chunking:
            return False

        try:
            doc = fitz.open(filepath)
            page_count = len(doc)
            doc.close()

            should_chunk = page_count >= self.chunk_threshold

            if should_chunk:
                logger.info(
                    f"→ {page_count} pages >= {self.chunk_threshold} threshold - will use chunking"
                )

            return should_chunk
        except Exception as e:
            logger.warning(f"Could not determine page count: {e} - will not chunk")
            return False

    def split_pdf_into_chunks(self, filepath):
        """
        Split PDF into chunks for memory-efficient processing.

        Creates temporary PDF chunks of specified size to process large documents
        without running out of memory.

        Args:
            filepath: Path to the PDF file to split

        Returns:
            tuple: (chunk_files list, temp_directory path)
                chunk_files: List of dictionaries with chunk metadata
                    (path, chunk_num, start_page, end_page, page_count)
                temp_directory: Path to temporary directory containing chunks
        """
        filepath = Path(filepath)

        # Create temporary directory for chunks
        temp_dir = tempfile.mkdtemp(prefix="pdf_chunks_")
        logger.info(f"Splitting PDF into chunks of {self.chunk_size} pages")
        logger.info(f"Temporary chunk directory: {temp_dir}")

        return self._split_with_pymupdf(filepath, temp_dir)

    def _split_with_pymupdf(self, filepath, temp_dir):
        """
        Split PDF using PyMuPDF (fast and reliable).

        Args:
            filepath: Path to the PDF file
            temp_dir: Temporary directory for storing chunks

        Returns:
            tuple: (chunk_files list, temp_directory path)
        """
        doc = fitz.open(str(filepath))
        total_pages = len(doc)

        chunk_files = []
        chunk_num = 0

        logger.info("Using PyMuPDF for fast PDF splitting...")

        for start_page in range(0, total_pages, self.chunk_size):
            end_page = min(start_page + self.chunk_size, total_pages)
            chunk_num += 1

            # Create new PDF for this chunk
            chunk_doc = fitz.open()
            chunk_doc.insert_pdf(doc, from_page=start_page, to_page=end_page - 1)

            # Save chunk
            chunk_filename = f"{filepath.stem}_chunk_{chunk_num:03d}.pdf"
            chunk_path = Path(temp_dir) / chunk_filename
            chunk_doc.save(str(chunk_path))
            chunk_doc.close()

            chunk_info = {
                "path": str(chunk_path),
                "chunk_num": chunk_num,
                "start_page": start_page + 1,  # 1-indexed
                "end_page": end_page,
                "page_count": end_page - start_page,
            }
            chunk_files.append(chunk_info)
            logger.info(f"  Created chunk {chunk_num}: pages {start_page+1}-{end_page}")

        doc.close()
        logger.info(f"Split into {len(chunk_files)} chunks")
        return chunk_files, temp_dir

    def parse_pdf_chunk(self, chunk_path, chunk_num, start_page):
        """
        Parse a single PDF chunk.

        Args:
            chunk_path: Path to the chunk PDF file
            chunk_num: Chunk number (for logging)
            start_page: Starting page number in the original document (1-indexed)

        Returns:
            dict: Dictionary containing:
                - result: Docling conversion result
                - toc: Table of contents string with adjusted page numbers
                - success: Boolean indicating success/failure
                - error: Error message if failed (optional)
                - chunk_path: Path to the chunk file
        """
        try:
            # Verify chunk file exists and get page count
            chunk_path_obj = Path(chunk_path)
            if not chunk_path_obj.exists():
                raise FileNotFoundError(f"Chunk file not found: {chunk_path}")

            # Verify it's actually a small chunk, not the full PDF
            chunk_doc = fitz.open(chunk_path)
            chunk_pages = len(chunk_doc)
            chunk_doc.close()

            logger.info(f"  Chunk file: {chunk_path_obj.name}")
            logger.info(
                f"  Chunk has {chunk_pages} pages (verifying it's a chunk, not full doc)"
            )

            # Convert chunk
            logger.info("  Converting chunk %d...", chunk_num)
            logger.info("  >>> Docling is processing: %s", chunk_path)
            result = self.converter.convert(source=chunk_path)
            logger.info("  ✓ Conversion complete")

            # Verify result page count matches chunk
            result_pages = len(result.document.pages)
            if result_pages != chunk_pages:
                logger.warning(
                    "  Page count mismatch: chunk has %d pages, result has %d pages",
                    chunk_pages,
                    result_pages,
                )
            else:
                logger.info("  ✓ Verified: processed %d pages from chunk", result_pages)

            # Apply hierarchical postprocessor
            logger.info("  Applying hierarchical postprocessor...")
            ResultPostprocessor(result, source=chunk_path).process()
            logger.info("  ✓ Postprocessing complete")

            # Generate TOC for this chunk (adjust page numbers)
            toc_lines = []
            for item, level in result.document.iterate_items():
                if hasattr(item, "text") and type(item).__name__ == "SectionHeaderItem":
                    indent = "  " * level
                    # Docling's item.level is 0-indexed, but markdown uses 1-indexed
                    # (level 0 = #, level 1 = ##, etc.)
                    # So we add 1 to match the markdown heading level
                    heading_level = getattr(item, "level", -1)
                    if heading_level != -1:
                        heading_level += 1  # Convert to markdown heading level
                    else:
                        heading_level = "?"

                    # Get page number from provenance and adjust for chunk offset
                    # NOTE: Docling's page_no is 1-indexed (first page = 1)
                    # For chunks: chunk page 1 = original page start_page
                    page_num = "?"
                    if hasattr(item, "prov") and item.prov:
                        for prov_item in item.prov:
                            if hasattr(prov_item, "page_no"):
                                # Adjust: chunk page_no 1 -> start_page
                                page_num = start_page + prov_item.page_no - 1
                                break

                    heading_text = item.text[:80]
                    toc_entry = (
                        f"{indent}[H{heading_level}] {heading_text} | page {page_num}"
                    )
                    toc_lines.append(toc_entry)

            toc_string = "\n".join(toc_lines)

            logger.info(
                f"  ✓ Generated TOC with {len(toc_lines)} headings for chunk {chunk_num}"
            )

            return {
                "result": result,
                "toc": toc_string,
                "success": True,
                "chunk_path": chunk_path,  # Keep track of chunk file
            }
        except Exception as e:
            logger.error(f"  ✗ Error parsing chunk {chunk_num}: {e}")
            return {
                "result": None,
                "toc": "",
                "success": False,
                "error": str(e),
            }

    def merge_chunk_results(
        self, chunk_results, output_folder, pdf_filename, chunk_files
    ):
        """
        Merge parsed chunk results into single output.

        Combines markdown files, TOC entries, and JSON documents from all chunks
        into unified output files.

        Args:
            chunk_results: List of dictionaries containing parse results for each chunk
            output_folder: Destination folder for merged output files
            pdf_filename: Base filename for output files (without extension)
            chunk_files: List of chunk metadata dictionaries

        Returns:
            tuple: (markdown_path, merged_toc) - paths/content of merged files
        """
        successful_chunks = sum(1 for r in chunk_results if r["success"])
        failed_chunks = len(chunk_results) - successful_chunks
        logger.info(
            f"Merging {successful_chunks}/{len(chunk_results)} successful chunks..."
        )
        if failed_chunks > 0:
            logger.warning(
                f"  {failed_chunks} chunk(s) failed and will be skipped in the output"
            )

        # Merge markdown with native Docling page breaks
        markdown_path = Path(output_folder) / f"{pdf_filename}.md"
        logger.info("  Combining markdown files...")

        # Create images directory for chunks
        images_dir = Path(output_folder) / "images"
        images_dir.mkdir(exist_ok=True, parents=True)

        with open(markdown_path, "w", encoding="utf-8") as outfile:
            for i, chunk_result in enumerate(chunk_results, 1):
                chunk_info = chunk_files[i - 1]
                start_page = chunk_info["start_page"]
                end_page = chunk_info["end_page"]

                if not chunk_result["success"]:
                    error_msg = chunk_result.get("error", "Unknown error")
                    pages = f"pages {start_page}-{end_page}"
                    # Add page separator before error message
                    if i > 1 or start_page > 1:
                        outfile.write(f"\n\n------- Page {start_page} -------\n\n")
                    outfile.write(
                        f"\n<!-- CHUNK {i} ({pages}) - "
                        f"PARSING FAILED: {error_msg} -->\n\n"
                    )
                    continue

                # Get markdown content using Docling's native page break support
                # Save chunk markdown with images to a temp file first
                result = chunk_result["result"]
                chunk_md_path = Path(output_folder) / f"_chunk_{i}_temp.md"
                result.document.save_as_markdown(
                    filename=chunk_md_path,
                    artifacts_dir=images_dir,
                    image_mode=ImageRefMode.REFERENCED,
                    page_break_placeholder=PAGE_SEPARATOR.strip(),
                )

                # Read the markdown content back
                with open(chunk_md_path, "r", encoding="utf-8") as chunk_file:
                    markdown_content = chunk_file.read()

                # Delete temp chunk file
                chunk_md_path.unlink()

                # Post-process chunk content: replace page breaks with numbered ones
                # Chunk pages start at 1, but need to be offset to actual document pages
                placeholder = PAGE_SEPARATOR.strip()
                parts = markdown_content.split(placeholder)

                # Write first part (page start_page of the document)
                # For the very first page of the document, write content then separator
                # For subsequent chunks, there's already a separator at the end of previous chunk
                outfile.write(parts[0])
                outfile.write(f"\n\n------- Page {start_page} -------\n\n")

                # Write subsequent parts with corrected page numbers
                for j, part in enumerate(parts[1:], 1):
                    page_num = start_page + j
                    outfile.write(part)
                    outfile.write(f"\n\n------- Page {page_num} -------\n\n")

        logger.info(f"  ✓ Merged markdown: {markdown_path}")

        # Merge TOC
        logger.info("  Combining TOC entries...")
        toc_lines = []
        for i, chunk_result in enumerate(chunk_results, 1):
            toc_content = chunk_result.get("toc", "")
            if toc_content:
                toc_lines.append(toc_content)

        merged_toc = "\n".join(toc_lines)

        # Save TOC
        toc_path = Path(output_folder) / "toc.txt"
        with open(toc_path, "w", encoding="utf-8") as f:
            f.write(merged_toc)
        logger.info(f"  ✓ Merged TOC: {toc_path}")

        # Print merged TOC to log
        if merged_toc:
            logger.info("\n--- Table of Contents (Merged from Chunks) ---")
            logger.info(merged_toc)
            logger.info("--- End of TOC ---\n")

        # Merge JSON (combine all page data)
        logger.info("  Combining JSON documents...")
        json_path = Path(output_folder) / f"{pdf_filename}.json"
        # Use the first successful result as base and extend pages from others
        merged_doc = None
        for chunk_result in chunk_results:
            if chunk_result["success"] and chunk_result["result"]:
                if merged_doc is None:
                    merged_doc = chunk_result["result"].document
                else:
                    # Append pages from this chunk
                    # Check if pages is a list or other structure
                    if hasattr(merged_doc, "pages") and hasattr(
                        chunk_result["result"].document, "pages"
                    ):
                        chunk_pages = chunk_result["result"].document.pages
                        # Handle both list and dict structures
                        if isinstance(merged_doc.pages, list) and isinstance(
                            chunk_pages, list
                        ):
                            merged_doc.pages.extend(chunk_pages)
                        elif isinstance(merged_doc.pages, dict) and isinstance(
                            chunk_pages, dict
                        ):
                            merged_doc.pages.update(chunk_pages)

        if merged_doc:
            merged_doc.save_as_json(json_path)
            logger.info(f"  ✓ Merged JSON: {json_path}")

        return str(markdown_path), merged_toc

    def cleanup_chunks(self, chunk_files, temp_dir):
        """
        Clean up temporary chunk files and directory.

        Args:
            chunk_files: List of chunk metadata dictionaries with file paths
            temp_dir: Path to temporary directory containing chunks
        """
        logger.info("Cleaning up chunk files...")

        for chunk_info in chunk_files:
            try:
                chunk_path = Path(chunk_info["path"])
                if chunk_path.exists():
                    chunk_path.unlink()
            except Exception as e:
                logger.warning(f"  Could not delete chunk: {e}")

        # Remove temp directory
        try:
            temp_dir = Path(temp_dir)
            if temp_dir.exists():
                temp_dir.rmdir()
                logger.info(f"✓ Removed temp directory: {temp_dir}")
        except Exception as e:
            logger.warning(f"  Could not remove temp directory: {e}")

    def _add_page_numbers_to_breaks(self, markdown_path, document):
        """
        Post-process markdown to add page numbers to page break placeholders.

        Analyzes the Docling document to determine page boundaries and updates
        generic page break markers with actual page numbers.

        Args:
            markdown_path: Path to the markdown file to process
            document: Docling document object with page information
        """
        try:
            # Read the markdown
            with open(markdown_path, "r", encoding="utf-8") as f:
                content = f.read()

            # Build a map of content positions to page numbers
            # We'll track which page each section of content belongs to
            page_content_map = []  # List of (page_num, text_snippet)

            for item, level in document.iterate_items():
                if hasattr(item, "text") and hasattr(item, "prov") and item.prov:
                    text = item.text.strip()
                    if text and len(text) > 20:  # Only substantial text
                        for prov_item in item.prov:
                            if hasattr(prov_item, "page_no"):
                                page_num = prov_item.page_no
                                page_content_map.append((page_num, text[:100]))
                                break

            # Split content by page break markers
            placeholder = PAGE_SEPARATOR.strip()
            parts = content.split(placeholder)

            if len(parts) <= 1:
                # No page breaks found
                return

            # First part is page 1
            result_parts = [parts[0], "\n\n------- Page 1 -------\n\n"]
            current_page = 1

            # For each subsequent part, determine its page number
            for i, part in enumerate(parts[1:], 1):
                # Find content in this part to determine page number
                part_preview = part.strip()[:200]  # First 200 chars

                # Try to match with our page content map
                matched_page = current_page + 1  # Default: increment
                for page_num, text_snippet in page_content_map:
                    if text_snippet in part_preview and page_num > current_page:
                        matched_page = page_num
                        break

                current_page = matched_page
                # Add the part then the break with page number
                result_parts.append(part)
                result_parts.append(f"\n\n------- Page {current_page} -------\n\n")

            # Write back
            new_content = "".join(result_parts)
            with open(markdown_path, "w", encoding="utf-8") as f:
                f.write(new_content)

            logger.debug(f"Added page numbers to {len(parts) - 1} page breaks")

        except Exception as e:
            logger.warning(f"Could not add page numbers to breaks: {e}")

    def detect_language(self, filepath):
        """
        Detect document language from first few pages of text.

        Uses langdetect library to identify the primary language in the document.
        Extracts text from the first 10 pages (or middle pages if beginning has
        insufficient text) to perform detection.

        Args:
            filepath: Path to the PDF file

        Returns:
            str: ISO language code (e.g., 'en', 'fr', 'es') or 'Unknown' if detection fails
        """
        logger.info("Detecting document language...")
        try:
            sample_text = ""

            # Extract text from first few pages (up to 10 pages to ensure enough text)
            doc = fitz.open(filepath)
            total_pages = len(doc)
            pages_to_check = min(10, total_pages)
            logger.debug(
                f"Extracting text from first {pages_to_check} pages (using PyMuPDF)"
            )

            for i in range(pages_to_check):
                try:
                    page_text = doc[i].get_text()
                    if page_text:
                        sample_text += page_text + " "
                        # Stop early if we have enough text (2000 chars for better accuracy)
                        if len(sample_text) > 2000:
                            break
                except Exception:
                    continue

            # Fallback: if not enough text from beginning, try middle pages
            sample_text = " ".join(sample_text.split())
            if len(sample_text) < 200 and total_pages > 10:
                logger.warning(
                    f"Not enough text from first pages (only {len(sample_text)} chars)"
                )
                logger.info("Trying 2 pages from middle of document...")
                sample_text = ""

                # Get 2 pages from the middle
                middle_start = total_pages // 2 - 1
                middle_pages = [middle_start, middle_start + 1]

                for page_num in middle_pages:
                    if 0 <= page_num < total_pages:
                        try:
                            page_text = doc[page_num].get_text()
                            if page_text:
                                sample_text += page_text + " "
                        except Exception:
                            continue

                logger.debug(
                    f"Extracted {len(sample_text)} chars from middle pages {middle_pages}"
                )

            doc.close()

            # Clean up text (remove extra whitespace)
            sample_text = " ".join(sample_text.split())
            text_length = len(sample_text)
            logger.debug(
                f"Final: Extracted {text_length} characters for language detection"
            )

            # Need at least 200 characters for reliable detection
            if text_length < 200:
                logger.warning(
                    f"Not enough text for language detection (only {text_length} chars, need 200+)"
                )
                return "Unknown"

            # Detect language using langdetect
            # langdetect is probabilistic, so we'll detect multiple times and use consensus
            from langdetect import detect_langs

            # Get top language probabilities
            lang_results = detect_langs(sample_text)

            if not lang_results:
                logger.warning("Language detection returned no results")
                return "Unknown"

            # Get the most confident result
            top_lang = lang_results[0]
            confidence = top_lang.prob
            lang_code = top_lang.lang

            # Log all detected languages for debugging
            langs_str = ", ".join(
                [f"{lang.lang}({lang.prob:.2f})" for lang in lang_results[:3]]
            )
            logger.info(f"✓ Language detection: {langs_str}")

            # Only accept if confidence is at least 0.8 (80%)
            if confidence < 0.8:
                logger.warning(f"Low confidence ({confidence:.2f}) - language unclear")
                # Return the best guess but log it
                logger.info(f"Best guess: {lang_code} (low confidence)")

            return lang_code

        except Exception as e:
            logger.warning(f"Language detection failed: {e}")
            return "Unknown"

    def parse_pdf_with_docling(self, filepath, output_folder):
        """
        Parse document (PDF or DOCX) using Docling Python API with hierarchical postprocessor.

        Performs complete document parsing including:
        - Language detection
        - Text extraction with structure preservation
        - Table of contents generation with hierarchical heading detection
        - Page count and word count statistics
        - Image extraction

        Args:
            filepath: Path to the PDF or DOCX file
            output_folder: Destination folder for parsed output

        Returns:
            tuple: (markdown_path, toc_string, page_count, word_count, language, file_format)
                Returns (None, None, None, None, None, None) if parsing fails
        """
        logger.info(f"Parsing: {filepath}")
        logger.info(f"Output folder: {output_folder}")

        # Detect file format
        file_ext = Path(filepath).suffix.lower()
        file_format = file_ext[1:] if file_ext else "unknown"  # Remove leading dot
        logger.info(f"File format: {file_format.upper()}")

        language = "Unknown"  # Default value in case detection fails

        try:
            # Detect language first
            language = self.detect_language(filepath)

            # Get page count (only for PDFs, DOCX doesn't have traditional pages)
            page_count = None
            if file_format == "pdf":
                try:
                    doc = fitz.open(filepath)
                    page_count = len(doc)
                    doc.close()
                    logger.info(f"Document has {page_count} pages")
                except Exception as e:
                    logger.warning(f"Could not determine page count: {e}")
            else:
                logger.info("Document is not a PDF (no page count available)")

            # Check if PDF should be chunked (only for large PDFs, not for DOCX)
            if file_format == "pdf" and self.should_chunk_pdf(filepath):
                return self.parse_pdf_with_chunking(filepath, output_folder)

            # Standard non-chunked parsing
            logger.info("Converting document...")
            result = self.converter.convert(source=filepath)
            logger.info("✓ Conversion complete")

            # Apply hierarchical postprocessor
            logger.info("Applying hierarchical postprocessor...")
            ResultPostprocessor(result, source=filepath).process()
            logger.info("✓ Hierarchical postprocessing complete")

            # Generate TOC with page numbers
            logger.info("Generating Table of Contents...")
            toc_lines = []
            for item, level in result.document.iterate_items():
                if hasattr(item, "text") and type(item).__name__ == "SectionHeaderItem":
                    indent = "  " * level
                    # Docling's item.level is 0-indexed, but markdown uses 1-indexed
                    # (level 0 = #, level 1 = ##, etc.)
                    # So we add 1 to match the markdown heading level
                    heading_level = getattr(item, "level", -1)
                    if heading_level != -1:
                        heading_level += 1  # Convert to markdown heading level
                    else:
                        heading_level = "?"

                    # Get page number from provenance
                    # NOTE: Docling's page_no is already 1-indexed (first page = 1)
                    page_num = "?"
                    if hasattr(item, "prov") and item.prov:
                        for prov_item in item.prov:
                            if hasattr(prov_item, "page_no"):
                                page_num = prov_item.page_no  # Already 1-indexed
                                break

                    heading_text = item.text[:80]
                    toc_entry = (
                        f"{indent}[H{heading_level}] {heading_text} | page {page_num}"
                    )
                    toc_lines.append(toc_entry)

            toc_string = "\n".join(toc_lines)
            logger.info(f"✓ Generated TOC with {len(toc_lines)} headings")

            # Print TOC to log
            if toc_lines:
                logger.info("\n--- Table of Contents ---")
                logger.info(toc_string)
                logger.info("--- End of TOC ---\n")

            # Save markdown with images and page breaks
            pdf_filename = Path(filepath).stem
            markdown_path = Path(output_folder) / f"{pdf_filename}.md"
            logger.info(f"Saving markdown to: {markdown_path}")

            # Use save_as_markdown to properly save images to subfolder
            result.document.save_as_markdown(
                filename=markdown_path,
                artifacts_dir=Path(output_folder) / "images",
                image_mode=ImageRefMode.REFERENCED,
                page_break_placeholder=PAGE_SEPARATOR.strip(),
            )

            # Post-process: Add page numbers to page breaks
            self._add_page_numbers_to_breaks(markdown_path, result.document)

            logger.info("✓ Markdown saved with page breaks and images")

            # Save JSON
            json_path = Path(output_folder) / f"{pdf_filename}.json"
            result.document.save_as_json(json_path)
            logger.debug(f"JSON saved to: {json_path}")

            # Save TOC
            toc_path = Path(output_folder) / "toc.txt"
            with open(toc_path, "w", encoding="utf-8") as f:
                f.write(toc_string)
            logger.info(f"✓ TOC saved to: {toc_path}")

            # Create symlink to original PDF
            self.create_pdf_symlink(filepath, output_folder)

            # Calculate word count from markdown
            with open(markdown_path, "r", encoding="utf-8") as f:
                markdown_content = f.read()
            word_count = len(markdown_content.split())

            logger.info(
                "✓ Document statistics: %d pages, %d words",
                page_count if page_count else 0,
                word_count,
            )
            logger.info("✓ Parsing successful")
            return (
                str(markdown_path),
                toc_string,
                page_count,
                word_count,
                language,
                file_format,
            )

        except FileNotFoundError as e:
            logger.error(f"✗ File not found: {e}")
            return None, None, None, None, None, None
        except Exception as e:
            logger.error(f"✗ Error parsing PDF: {e}")
            import traceback

            logger.error(traceback.format_exc())
            return None, None, None, None, None, None

    def parse_pdf_with_chunking(self, filepath, output_folder):
        """
        Parse large PDF using chunking approach for memory efficiency.

        Splits large PDFs into smaller chunks, processes each separately, and
        merges the results. This prevents memory issues with very large documents.

        Args:
            filepath: Path to the PDF file
            output_folder: Destination folder for parsed output

        Returns:
            tuple: (markdown_path, toc_string, total_pages, word_count, language, file_format)
                Returns (None, None, None, None, None, None) if parsing fails
        """
        logger.info("=" * 60)
        logger.info("USING CHUNKED PARSING (Memory-Efficient Mode)")
        logger.info("=" * 60)

        # File format detection (chunking only used for PDFs)
        file_ext = Path(filepath).suffix.lower()
        file_format = file_ext[1:] if file_ext else "unknown"

        chunk_files = None
        temp_dir = None

        try:
            # Detect language first
            language = self.detect_language(filepath)
            # Split PDF into chunks
            chunk_files, temp_dir = self.split_pdf_into_chunks(filepath)

            logger.info("")
            logger.info("=" * 60)
            logger.info(f"PARSING {len(chunk_files)} CHUNKS")
            logger.info("=" * 60)

            # Parse each chunk
            chunk_results = []
            for i, chunk_info in enumerate(chunk_files, 1):
                pages_range = f"{chunk_info['start_page']}-{chunk_info['end_page']}"
                logger.info(
                    "\n[Chunk %d/%d] Processing pages %s...",
                    i,
                    len(chunk_files),
                    pages_range,
                )

                # Set timeout alarm
                timeout_seconds = self.chunk_timeout

                def timeout_handler(signum, frame):
                    raise TimeoutError(
                        f"Chunk {i} processing exceeded {timeout_seconds} seconds"
                    )

                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(timeout_seconds)

                try:
                    chunk_result = self.parse_pdf_chunk(
                        chunk_info["path"],
                        chunk_info["chunk_num"],
                        chunk_info["start_page"],
                    )
                    signal.alarm(0)  # Cancel alarm on success
                except TimeoutError as e:
                    signal.alarm(0)  # Cancel alarm
                    logger.error(f"✗ Chunk {i} TIMEOUT: {e}")
                    logger.warning(
                        f"  Skipping chunk {i} and continuing with remaining chunks..."
                    )
                    chunk_result = {
                        "result": None,
                        "toc": "",
                        "success": False,
                        "error": f"Timeout after {timeout_seconds} seconds",
                    }
                except Exception as e:
                    signal.alarm(0)  # Cancel alarm
                    logger.error(f"✗ Chunk {i} ERROR: {e}")
                    chunk_result = {
                        "result": None,
                        "toc": "",
                        "success": False,
                        "error": str(e),
                    }

                chunk_results.append(chunk_result)

                if chunk_result["success"]:
                    logger.info(f"✓ Chunk {i} parsed successfully")
                else:
                    logger.error(
                        f"✗ Chunk {i} failed: {chunk_result.get('error', 'Unknown error')}"
                    )

            logger.info("")
            logger.info("=" * 60)
            logger.info("MERGING CHUNK RESULTS")
            logger.info("=" * 60)

            # Merge results
            pdf_filename = Path(filepath).stem
            markdown_path, toc_string = self.merge_chunk_results(
                chunk_results, output_folder, pdf_filename, chunk_files
            )

            # Calculate page count and word count from merged markdown
            total_pages = sum(
                chunk_info["end_page"] - chunk_info["start_page"] + 1
                for chunk_info in chunk_files
            )
            word_count = 0
            if markdown_path and Path(markdown_path).exists():
                with open(markdown_path, "r", encoding="utf-8") as f:
                    merged_markdown = f.read()
                word_count = len(merged_markdown.split())

            logger.info(
                "✓ Total document statistics: %d pages, %d words",
                total_pages,
                word_count,
            )

            # Create symlink to original PDF
            self.create_pdf_symlink(filepath, output_folder)

            # Cleanup temporary files
            self.cleanup_chunks(chunk_files, temp_dir)

            logger.info("=" * 60)
            logger.info("✓ CHUNKED PARSING COMPLETE")
            logger.info("=" * 60)

            return (
                markdown_path,
                toc_string,
                total_pages,
                word_count,
                language,
                file_format,
            )

        except FileNotFoundError as e:
            logger.error(f"✗ File not found: {e}")
            return None, None, None, None, None, None
        except Exception as e:
            logger.error(f"✗ Error in chunked parsing: {e}")
            import traceback

            logger.error(traceback.format_exc())

            # Cleanup on error
            if chunk_files and temp_dir:
                try:
                    self.cleanup_chunks(chunk_files, temp_dir)
                except Exception:  # noqa: E722
                    pass

            return None, None, None, None, None, None

    def parse_pdf_with_protection(self, filepath, output_folder, timeout=900):
        """
        Parse PDF with subprocess protection to handle OOM kills and crashes.

        Runs the parsing operation in an isolated subprocess with timeout protection.
        This prevents Out-Of-Memory crashes from affecting the main process.

        Args:
            filepath: Path to PDF file
            output_folder: Output directory
            timeout: Timeout in seconds (default 900 = 15 minutes)

        Returns:
            tuple: Same as parse_pdf_with_docling:
                (markdown_path, toc_string, page_count, word_count,
                language, file_format)
                Returns (None, None, None, None, None, None) if fails
        """
        logger.info("=" * 60)
        logger.info("USING SUBPROCESS PROTECTION (OOM-Safe Mode)")
        logger.info("=" * 60)

        # Prepare parser arguments for subprocess
        parser_args = {
            "metadata_path": self.metadata_path,
            "output_dir": self.output_dir,
            "pipeline": self.pipeline,
            "ocr_engine": self.ocr_engine,
            "ocr_lang": self.ocr_lang,
            "pdf_backend": self.pdf_backend,
            "table_mode": self.table_mode,
            "no_ocr": self.no_ocr,
            "enrich_picture_description": self.enrich_picture_description,
            "images_scale": self.images_scale,
            "enable_chunking": self.enable_chunking,
            "chunk_size": self.chunk_size,
            "chunk_threshold": self.chunk_threshold,
            "chunk_timeout": self.chunk_timeout,
        }

        # Create a queue for inter-process communication
        result_queue = multiprocessing.Queue()

        # Create and start the worker process
        process = multiprocessing.Process(
            target=_parse_pdf_worker,
            args=(filepath, output_folder, parser_args, result_queue),
        )

        try:
            logger.info(f"Starting parsing subprocess (timeout: {timeout}s)...")
            process.start()
            process.join(timeout=timeout)

            if process.is_alive():
                # Process exceeded timeout
                logger.error(
                    f"✗ Process exceeded timeout of {timeout}s - terminating..."
                )
                process.terminate()
                process.join(timeout=10)

                if process.is_alive():
                    # Force kill if it doesn't terminate gracefully
                    logger.error("✗ Process didn't terminate gracefully - killing...")
                    process.kill()
                    process.join()

                logger.error(f"✗ Parsing timeout after {timeout}s")
                return None, None, None, None, None, None

            # Check exit code
            if process.exitcode != 0:
                logger.error(f"✗ Process exited with code {process.exitcode}")

                # Try to get error from queue
                if not result_queue.empty():
                    result = result_queue.get(timeout=1)
                    if not result.get("success"):
                        error_msg = result.get("error", "Unknown error")
                        logger.error(f"✗ Error details: {error_msg}")
                        if result.get("traceback"):
                            logger.error(f"Traceback:\n{result.get('traceback')}")
                else:
                    # Process was likely killed (OOM, segfault, etc.)
                    if process.exitcode == -9:
                        logger.error(
                            "✗ Process was KILLED by system (Out Of Memory - OOM)"
                        )
                        logger.error(
                            "✗ Document is too large/complex for available memory"
                        )
                    elif process.exitcode == -11:
                        logger.error("✗ Process crashed (segmentation fault)")
                    else:
                        logger.error(f"✗ Process failed with signal {process.exitcode}")

                return None, None, None, None, None, None

            # Process completed successfully - get result
            if result_queue.empty():
                logger.error("✗ Process completed but no result available")
                return None, None, None, None, None, None

            result = result_queue.get(timeout=5)

            if result.get("success"):
                logger.info("✓ Parsing completed successfully")
                return result["result"]
            else:
                error_msg = result.get("error", "Unknown error")
                logger.error(f"✗ Parsing failed: {error_msg}")
                if result.get("traceback"):
                    logger.error(f"Traceback:\n{result.get('traceback')}")
                return None, None, None, None, None, None

        except Exception as e:
            logger.error(f"✗ Unexpected error in subprocess management: {e}")
            import traceback

            logger.error(traceback.format_exc())

            # Cleanup process if still running
            if process.is_alive():
                process.terminate()
                process.join(timeout=5)
                if process.is_alive():
                    process.kill()
                    process.join()

            return None, None, None, None, None, None

    def _ensure_columns_exist(self, ws, wb):
        """
        Ensure all required columns exist in the metadata sheet.

        Adds any missing columns to the Excel worksheet header row.

        Args:
            ws: Excel worksheet
            wb: Excel workbook
        """
        headers = [cell.value for cell in ws[1]]
        columns_added = False

        required_columns = [
            "Parsed Folder",
            "Parsed Markdown Path",
            "TOC",
            "Page Count",
            "Word Count",
            "Language",
            "File Format",
            "Parsed Error",
        ]

        for column_name in required_columns:
            if column_name not in headers:
                col_idx = len(headers) + 1
                ws.cell(1, col_idx, column_name)
                headers.append(column_name)
                logger.info(f"Added '{column_name}' column to metadata")
                columns_added = True

        # Save once if any columns were added
        if columns_added:
            wb.save(self.metadata_path)
            logger.info("✓ Metadata columns updated and saved")

    def update_metadata(
        self,
        ws,
        row_idx,
        parsed_folder,
        markdown_path,
        toc_string,
        page_count=None,
        word_count=None,
        parsing_error=None,
        language=None,
        file_format=None,
    ):
        """
        Update metadata with parsing results.

        Updates Excel worksheet with TOC, page count, word count, language,
        file format, and errors.

        Args:
            ws: Excel worksheet
            row_idx: Row index to update
            parsed_folder: Path to the output folder
            markdown_path: Path to the generated markdown file
            toc_string: Table of contents string
            page_count: Number of pages in document
            word_count: Number of words in document
            parsing_error: Error message if parsing failed
            language: Detected language code
            file_format: File format (PDF, DOCX)
        """
        headers = [cell.value for cell in ws[1]]

        # Get column indices (columns should already exist from _ensure_columns_exist)
        parsed_folder_col = headers.index("Parsed Folder") + 1
        parsed_markdown_col = headers.index("Parsed Markdown Path") + 1
        toc_col = headers.index("TOC") + 1
        page_count_col = headers.index("Page Count") + 1
        word_count_col = headers.index("Word Count") + 1
        language_col = headers.index("Language") + 1
        file_format_col = headers.index("File Format") + 1
        parsing_error_col = headers.index("Parsed Error") + 1

        # Update values
        ws.cell(row_idx, parsed_folder_col, parsed_folder if parsed_folder else "")
        ws.cell(row_idx, parsed_markdown_col, markdown_path if markdown_path else "")
        ws.cell(row_idx, toc_col, toc_string if toc_string else "")
        ws.cell(row_idx, page_count_col, page_count if page_count else "")
        ws.cell(row_idx, word_count_col, word_count if word_count else "")
        ws.cell(row_idx, language_col, language if language else "Unknown")
        ws.cell(row_idx, file_format_col, file_format.upper() if file_format else "")
        # Only write parsing_error if it's not None and not empty string
        ws.cell(row_idx, parsing_error_col, parsing_error if parsing_error else "")

    def process_document(self, filepath):
        """
        Process a single document and return the result.
        Wrapper around internal parsing logic.
        """
        try:
            # Create output folder
            output_folder = self.create_output_folder(filepath)

            # Determine if we should use chunking
            use_chunking = self.enable_chunking and self.should_chunk_pdf(filepath)

            if use_chunking:
                logger.info(f"Processing with chunking: {filepath}")
                markdown_path, toc, pages, words, lang, fmt = (
                    self.parse_pdf_with_chunking(filepath, output_folder)
                )
            else:
                logger.info(f"Processing standard: {filepath}")
                markdown_path, toc, pages, words, lang, fmt = (
                    self.parse_pdf_with_docling(filepath, output_folder)
                )

            if markdown_path:
                return {
                    "success": True,
                    "output_dir": output_folder,
                    "toc": toc,
                    "pages": pages,
                    "words": words,
                    "language": lang,
                    "format": fmt,
                }
            else:
                return {"success": False, "error": "Parsing returned None"}

        except Exception as e:
            logger.error(f"Error processing document {filepath}: {e}")
            return {"success": False, "error": str(e)}

    def parse_all(self, year=None, agency=None, force_parse=False):
        """
        Parse all PDFs based on metadata and filters.

        Main entry point for batch processing documents. Reads metadata from Excel,
        filters documents based on criteria, and processes each one.

        Args:
            year: Optional year filter to limit parsing to specific year
            agency: Optional agency filter to limit parsing to specific agency
            force_parse: If True, re-parse documents even if already parsed
        """
        logger.info("=" * 60)
        logger.info("Starting PDF parsing with Docling")
        logger.info("=" * 60)

        wb, ws = self.load_metadata()
        if not wb or not ws:
            return

        # Ensure all required columns exist upfront
        self._ensure_columns_exist(ws, wb)

        rows_to_parse = self.get_rows_to_parse(ws, year, agency, force_parse)

        if not rows_to_parse:
            logger.warning("No documents to parse")
            return

        parsed_count = 0
        failed_count = 0

        for idx, row_data in enumerate(rows_to_parse, 1):
            row_idx = row_data["row_idx"]
            filepath = row_data["filepath"]

            logger.info(f"\n[{idx}/{len(rows_to_parse)}] Processing document...")
            logger.info(f"  Agency: {row_data.get('agency', 'N/A')}")
            logger.info(f"  Year: {row_data.get('year', 'N/A')}")
            logger.info(f"  File: {filepath}")

            # Create output folder
            output_folder = self.create_output_folder(filepath)

            # Check if markdown already exists in output folder (unless force_parse)
            if not force_parse:
                pdf_filename = Path(filepath).stem
                expected_markdown = Path(output_folder) / f"{pdf_filename}.md"

                if expected_markdown.exists():
                    logger.info(f"  ✓ Markdown already exists: {expected_markdown}")

                    # Read existing markdown to get word count
                    try:
                        with open(expected_markdown, "r", encoding="utf-8") as f:
                            markdown_content = f.read()
                        word_count = len(markdown_content.split())

                        # Try to get page count from markdown
                        import re

                        page_matches = re.findall(
                            r"------- Page (\d+) -------", markdown_content
                        )
                        page_count = (
                            max([int(p) for p in page_matches])
                            if page_matches
                            else None
                        )

                        # Read TOC if it exists
                        toc_path = Path(output_folder) / "toc.txt"
                        toc_string = ""
                        if toc_path.exists():
                            with open(toc_path, "r", encoding="utf-8") as f:
                                toc_string = f.read()

                        # Language would be unknown since we're not re-parsing
                        language = "Unknown"

                        # Detect file format from filepath
                        file_ext = Path(filepath).suffix.lower()
                        file_format = file_ext[1:] if file_ext else "unknown"

                        # Update metadata with existing parse results
                        self.update_metadata(
                            ws,
                            row_idx,
                            output_folder,
                            str(expected_markdown),
                            toc_string,
                            page_count,
                            word_count,
                            None,  # No error
                            language,
                            file_format,
                        )

                        parsed_count += 1
                        logger.info("  ✓ Metadata updated with existing parse results")

                        # Save after each document
                        wb.save(self.metadata_path)
                        logger.debug("  Metadata updated in Excel")
                        continue

                    except Exception as e:
                        logger.warning(f"  Could not read existing markdown: {e}")
                        # Fall through to re-parse

            # Check if file exists
            if not os.path.exists(filepath):
                error_msg = "File not found"
                logger.error(f"  ✗ {error_msg}: {filepath}")
                self.update_metadata(
                    ws, row_idx, "", "", "", None, None, error_msg, None, None
                )
                failed_count += 1
                wb.save(self.metadata_path)
                continue

            # Parse with subprocess protection (handles OOM kills and crashes)
            result = self.parse_pdf_with_protection(
                filepath, output_folder, timeout=900
            )
            (
                markdown_path,
                toc_string,
                page_count,
                word_count,
                language,
                file_format,
            ) = result

            # Determine parsing error status - only set error if parsing actually failed
            parsing_error = None
            if not markdown_path:
                parsing_error = (
                    "Parsing failed - likely Out Of Memory (OOM) - "
                    "document too large/complex"
                )

            # Update metadata
            self.update_metadata(
                ws,
                row_idx,
                output_folder,
                markdown_path,
                toc_string,
                page_count,
                word_count,
                parsing_error,
                language,
                file_format,
            )

            if markdown_path:
                parsed_count += 1
                logger.info(f"  ✓ Success - Markdown saved to: {markdown_path}")
            else:
                failed_count += 1
                logger.error("  ✗ Failed to parse document")

            # Save after each document
            wb.save(self.metadata_path)
            logger.debug("  Metadata updated in Excel")

        logger.info("\n" + "=" * 60)
        logger.info("Parsing complete!")
        logger.info(f"Successfully parsed: {parsed_count}")
        logger.info(f"Failed: {failed_count}")
        logger.info(f"Total processed: {len(rows_to_parse)}")
        logger.info(f"Metadata updated: {self.metadata_path}")
        logger.info("=" * 60)


def main():
    parser = DoclingParser()

    # Get documents to parse from Qdrant
    logger.info("Fetching 'downloaded' documents from Qdrant...")
    docs_to_parse = db.get_documents_by_status("downloaded")

    if not docs_to_parse:
        logger.info("No documents found with status 'downloaded'.")
        return

    logger.info(f"Found {len(docs_to_parse)} documents to parse.")

    for doc in docs_to_parse:
        doc_id = generate_doc_id(doc.get("url") or doc.get("filepath"))
        filepath = doc.get("filepath")

        if not filepath or not os.path.exists(filepath):
            logger.warning(f"File not found for {doc.get('title')}: {filepath}")
            db.update_document(
                doc_id, {"status": "parse_failed", "processing_error": "File not found"}
            )
            continue

        logger.info(f"Processing: {doc.get('title')}")

        try:
            # Parse the document
            result = parser.process_document(filepath)

            if result["success"]:
                # Update Qdrant
                # result['output_dir'] is the folder containing the parsed JSON
                db.update_document(
                    doc_id,
                    {
                        "status": "parsed",
                        "parsed_folder": str(result["output_dir"]),
                        "toc": result.get("toc", ""),
                    },
                )
                logger.info(f"Successfully parsed {doc.get('title')}")
            else:
                db.update_document(
                    doc_id,
                    {
                        "status": "parse_failed",
                        "processing_error": result.get("error", "Unknown error"),
                    },
                )
                logger.error(f"Failed to parse {doc.get('title')}")

        except Exception as e:
            logger.error(f"Exception parsing {doc.get('title')}: {e}")
            db.update_document(
                doc_id, {"status": "parse_failed", "processing_error": str(e)}
            )


if __name__ == "__main__":
    main()
