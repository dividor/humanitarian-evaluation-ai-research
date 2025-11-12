"""
summarize.py - Generate summaries for parsed documents using sentence transformers

Creates three types of summaries:
1. Centroid-based: Most representative sentences from entire document
2. Summary section-based: Extract from identified summary sections (Executive Summary, etc.)
3. Query-based: Sentences relevant to specific questions

Updates metadata Excel file with summary columns.
"""

import argparse
import logging
import os
import re
import sys
from pathlib import Path

import nltk
import numpy as np
import openpyxl
import requests
from deep_translator import GoogleTranslator
from dotenv import load_dotenv
from nltk.tokenize import sent_tokenize
from sentence_transformers import SentenceTransformer, util

# Load environment variables
load_dotenv()

# Ensure NLTK data is available
try:
    nltk.data.find("tokenizers/punkt")
except LookupError:
    print("Downloading NLTK punkt tokenizer...")
    nltk.download("punkt", quiet=True)

try:
    nltk.data.find("tokenizers/punkt_tab")
except LookupError:
    print("Downloading NLTK punkt_tab tokenizer...")
    nltk.download("punkt_tab", quiet=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Used for identifying summary sections and centroid sentences for extractive summary
EMBEDDING_MODEL = "all-MiniLM-L6-v2"

# LLM Model Configuration
# ALL models use HuggingFace Router API via direct requests
# Requires HUGGINGFACE_API_KEY or HF_TOKEN environment variable
#
# Configuration parameters:
# - model: Model identifier with provider suffix (e.g., "model-name:provider")
# - provider: Provider name (hf-inference, featherless-ai, novita, etc.)
# - max_context_chars: Maximum INPUT length the model can process in one call
# - max_tokens: Maximum OUTPUT length for final summary generation
# - chunk_overlap: Character overlap between chunks (for context continuity)
# - chunk_tokens_ratio: Ratio for intermediate chunk summaries (default: 0.5)
#   * chunk_tokens = max_tokens * chunk_tokens_ratio
#   * Smaller ratio (e.g., 0.4) = shorter chunk summaries, more room for combining
#   * Larger ratio (e.g., 0.6) = longer chunk summaries, may require more reduction
#   * Recommended: 0.5 provides good balance across all models
#
# Uncomment one model configuration to use it

# Option 1: BART - Uses Router API hf-inference endpoint
# LLM_MODEL = {
#    "model": "facebook/bart-large-cnn",
#    "provider": "hf-inference",
#    "max_context_chars": 4000,
#    "max_tokens": 1024,
#    "chunk_overlap": 400,
#    "chunk_tokens_ratio": 0.5,  # chunk_tokens = 1024 * 0.5 = 512
# }

# Option 2: Mistral (larger context, better quality) - Router API with featherless-ai provider
# LLM_MODEL = {
#    "model": "mistralai/Mistral-7B-Instruct-v0.2:featherless-ai",
#    "provider": "featherless-ai",
#    "max_context_chars": 29000,
#    "max_tokens": 2000,
#    "chunk_overlap": 800,
#    "chunk_tokens_ratio": 0.5,  # chunk_tokens = 1500 * 0.5 = 750
# }

# Option 3: Llama - Router API with novita provider
LLM_MODEL = {
    "model": "meta-llama/Llama-3.1-8B-Instruct:novita",
    "provider": "novita",
    "max_context_chars": 29000,
    "max_tokens": 2000,
    "chunk_overlap": 800,
    "chunk_tokens_ratio": 0.5,  # chunk_tokens = 1024 * 0.5 = 512
}

# Model temperature.
# 0.0 = most deterministic, 1.0 = most random
LLM_TEMPERATURE = 0.1

# Used as fallback or when extractive method is centroid.
NUM_CENTROID_SENTENCES = 30

# Extractive summary method for LLM input
# Controls which extractive summary to use as input for LLM summarization
# Options:
# - "summary_sections": Use TOC-based summary sections (Executive Summary, etc.)
#   Falls back to centroid if no good TOC found
# - "centroid": Always use centroid-based extractive summary
#   Uses top NUM_CENTROID_SENTENCES sentences by semantic similarity
EXTRACTIVE_SUMMARY_METHOD = "centroid"  # or "centroid"

# LLM Prompts
# LLM_REDUCTION_PROMPT: Used during MAP phase to summarize each chunk
LLM_REDUCTION_PROMPT = """
You are an expert research summarizer. Summarize the following text with a
focus on identifying its key *concepts*, *methods*, *findings*, and
*terminology*.

Organize your output as a structured summary to support taxonomy development.
Capture recurring or novel terms, thematic groupings, and conceptual
relationships.

Include both positive and negative findings or limitations. Do not assume
overall success unless explicitly supported.

Only include findings or claims explicitly stated in the source text. Do not
infer new themes.

TEXT:
<<< {document_text} >>>

OUTPUT FORMAT:
- **Summary:** [2-3 paragraph summary of the report]
- **Main Topics:** [List key subjects, domains, or disciplines covered,
  with brief descriptions]
- **Key Terms & Concepts:** [Extract important domain-specific terminology
  and emerging phrases, with brief definitions]
- **Methods / Approach:** [Summarize methods, frameworks, or analytical
  strategies mentioned, with brief descriptions]
- **Findings / Conclusions:** [Summarize main insights or takeaways, with
  brief descriptions]
- **Related Concepts / Themes:** [List secondary ideas, linked concepts, or
  cross-cutting themes, with brief descriptions]
"""

# LLM_SUMMARY_PROMPT: Used for final summary generation (single-pass or REDUCE phase)
# Used for chat-based models (Mistral, Llama) to guide summary generation
# BART doesn't use system prompts (pure summarization model)
LLM_SUMMARY_PROMPT = """
You are a research analyst consolidating multiple structured summaries into a
comprehensive overview. Your task is to write a report synopsis, identify
overarching themes, recurring terminology, and conceptual groupings to support
taxonomy or ontology creation.

Include both positive and negative findings or limitations. Do not assume
overall success unless explicitly supported.

Only include findings or claims explicitly stated in the source text. Do not
infer new themes.

Here are the interim summaries:
<<< {map_summaries} >>>

YOU MUST ADHERE TO THE EXACT FORMAT and TITLES:
- **Title:** [title of the report]
- **Summary:** [2-3 paragraph summary of the report]
- **Report Summary:** [2-3 paragraph summary of the report]
- **Topics:** [Summarize the main thematic clusters emerging across documents,
  with brief descriptions]
- **Core Concepts and Terms:** [List and briefly define the most important
  recurring or novel terms, with brief definitions]
- **Methodological Patterns:** [Summarize common approaches or frameworks
  across sources, with brief descriptions]
- **Key Conclusions:** [Highlight consensus findings and important divergences,
  with brief descriptions]
- **Emergent Taxonomy Proposal:** [Organize related terms or ideas
  hierarchically; e.g. Category → Subcategory → Concept, with brief
  descriptions]
- **Notable Gaps or Contradictions:** [Mention where the data diverges or
  lacks clarity, with brief descriptions]
"""

# Summarization settings
NUM_CENTROID_SENTENCES = 30


def clean_markdown_formatting(text):
    """
    Clean up markdown formatting issues in LLM-generated text.

    Removes bold formatting from markdown headings to ensure proper formatting
    (e.g., ## **Title** -> ## Title).

    Args:
        text: Markdown text to clean

    Returns:
        str: Cleaned markdown text
    """
    if not text:
        return text

    # Remove bold formatting from markdown headings
    # Match: ## **Title** or ## ** Title ** and convert to ## Title
    text = re.sub(
        r"^(#{1,6})\s*\*\*\s*(.+?)\s*\*\*\s*$", r"\1 \2", text, flags=re.MULTILINE
    )

    return text


def get_model_type(model_config):
    """
    Detect model type based on model name.

    Identifies whether the configured LLM model is BART, Mistral, Llama, or
    a generic chat model. This determines which API format to use.

    Args:
        model_config: Model configuration dictionary or model name string

    Returns:
        str: Model type ('bart', 'mistral', 'llama', or 'chat')
    """
    model_name = (
        model_config["model"] if isinstance(model_config, dict) else model_config
    )
    if "bart" in model_name.lower():
        return "bart"
    elif "mistral" in model_name.lower():
        return "mistral"
    elif "llama" in model_name.lower():
        return "llama"
    else:
        # Default to chat completion API
        return "chat"


class DocumentSummarizer:
    """
    Document summarization using embeddings and LLM models.

    Provides multiple summarization methods:
    - Centroid-based extractive summarization using sentence embeddings
    - Summary section extraction from TOC (Executive Summary, etc.)
    - Abstractive summarization using LLM models via HuggingFace Router API
    - Language detection and translation to English

    Attributes:
        metadata_path: Path to Excel metadata file
        model: Sentence transformer model for embeddings
        hf_token: HuggingFace API token for LLM access
        model_type: Type of LLM model being used ('bart', 'mistral', 'llama', or 'chat')
    """

    def __init__(self, metadata_path="./data/pdf_metadata.xlsx"):
        """
        Initialize the DocumentSummarizer.

        Args:
            metadata_path: Path to the Excel metadata file containing document information
        """
        self.metadata_path = metadata_path
        self.model = None  # Sentence transformer model for embeddings
        self.hf_token = None  # HuggingFace API token
        self.model_type = get_model_type(LLM_MODEL)  # Detect model type

    def save_content_to_file(self, markdown_path, content, suffix):
        """
        Save content to a file in the same directory as the markdown file.

        Cleans the content (removes images, page separators, HTML comments) before
        saving to ensure clean text output.

        Args:
            markdown_path: Path to the source markdown file
            content: Content to save
            suffix: File suffix (e.g., 'key_content_sections', 'centroid', 'llm_summary')

        Returns:
            str: Relative path to the saved file (from data/ directory), or None if failed
        """
        if not content or not markdown_path:
            return None

        try:
            # Clean content before saving (remove images and page separators)
            cleaned_content = content

            # Remove markdown images: ![alt](path)
            cleaned_content = re.sub(r"!\[.*?\]\(.*?\)", "", cleaned_content)

            # Remove standalone image paths (lines that are just image paths)
            cleaned_content = re.sub(
                r"^\s*[^\s]+/images/image_[^\s]+\.(png|jpg|jpeg|gif)\s*$",
                "",
                cleaned_content,
                flags=re.MULTILINE,
            )

            # Remove page separators
            cleaned_content = re.sub(r"------- Page \d+ -------", "", cleaned_content)
            cleaned_content = re.sub(r"------- Page Break -------", "", cleaned_content)

            # Remove HTML comments
            cleaned_content = re.sub(
                r"<!--.*?-->", "", cleaned_content, flags=re.DOTALL
            )

            # Clean up excessive whitespace
            # For LLM summaries, preserve double newlines between chunks
            if suffix not in ["llm_chunked_summary", "llm_summary"]:
                # Remove lines that are only whitespace
                cleaned_content = re.sub(
                    r"^\s*$\n", "", cleaned_content, flags=re.MULTILINE
                )
                # Replace multiple newlines with double newline
                cleaned_content = re.sub(r"\n\s*\n\s*\n+", "\n\n", cleaned_content)
            else:
                # For LLM summaries, only collapse 3+ consecutive newlines to 2
                cleaned_content = re.sub(r"\n\s*\n\s*\n+", "\n\n", cleaned_content)

            cleaned_content = cleaned_content.strip()

            # Get the directory containing the markdown file
            md_path = Path(markdown_path)
            if not md_path.exists():
                logger.warning(f"Markdown file does not exist: {markdown_path}")
                return None

            parent_dir = md_path.parent

            # Create simple filename based on suffix only (without document name prefix)
            content_filename = f"{suffix}.txt"
            content_path = parent_dir / content_filename

            # Write cleaned content to file
            with open(content_path, "w", encoding="utf-8") as f:
                f.write(cleaned_content)

            # Return path with ./data/ prefix for consistency with markdown path format
            # Convert absolute path to relative from data/
            try:
                data_dir = Path("./data").resolve()
                relative_path = content_path.resolve().relative_to(data_dir)
                return f"./data/{relative_path}"
            except ValueError:
                # If path is not relative to data/, return absolute path
                return str(content_path)

        except Exception as e:
            logger.error(f"Failed to save content to file: {e}")
            return None

    def save_llm_summary_centralized(self, markdown_path, content, suffix):
        """
        Save LLM summary to both the document folder and centralized summaries folder.

        Creates two copies: one in the document's parsed folder and one in the
        centralized data/summaries/ directory for easy access.

        Args:
            markdown_path: Path to the source markdown file
            content: Summary content to save
            suffix: File suffix ('llm_summary' or 'llm_chunked_summary')

        Returns:
            tuple: (doc_folder_path, centralized_path) or (None, None) if failed
        """
        if not content or not markdown_path:
            return None, None

        try:
            # Save to document folder (existing behavior)
            doc_folder_path = self.save_content_to_file(markdown_path, content, suffix)

            # Also save to centralized data/summaries folder
            md_path = Path(markdown_path)
            if not md_path.exists():
                return doc_folder_path, None

            # Create centralized summaries directory
            summaries_dir = Path("./data/summaries")
            summaries_dir.mkdir(parents=True, exist_ok=True)

            # Create filename from the document's unique identifier
            # Extract organization/year/document_id from path
            # Format: data/parsed/ORG/YEAR/DOC_NAME_ID/file.md
            parts = md_path.parts
            if len(parts) >= 5 and parts[0] == "data" and parts[1] == "parsed":
                org = parts[2]
                year = parts[3]
                doc_folder = parts[4]  # e.g., "Document Title_123"

                # Create sanitized filename: ORG_YEAR_DOCFOLDER_suffix.txt
                centralized_filename = f"{org}_{year}_{doc_folder}_{suffix}.txt"
                centralized_path = summaries_dir / centralized_filename

                # Write to centralized location
                with open(centralized_path, "w", encoding="utf-8") as f:
                    f.write(content)

                logger.info(f"  ✓ Also saved to centralized: {centralized_path}")
                return doc_folder_path, str(centralized_path)
            else:
                logger.warning(
                    f"  Could not determine centralized path for: {markdown_path}"
                )
                return doc_folder_path, None

        except Exception as e:
            logger.error(f"Failed to save centralized summary: {e}")
            return doc_folder_path if "doc_folder_path" in locals() else None, None

    def load_content_from_file(self, file_path):
        """
        Load content from a file path.

        Args:
            file_path: Path to the content file (e.g., ./data/parsed/...)

        Returns:
            str: Content string, or None if failed
        """
        if not file_path or file_path in ["No", "No TOC", None, ""]:
            return None

        try:
            full_path = Path(file_path)

            if not full_path.exists():
                logger.warning(f"Content file does not exist: {full_path}")
                return None

            with open(full_path, "r", encoding="utf-8") as f:
                return f.read()

        except Exception as e:
            logger.error(f"Failed to load content from file: {e}")
            return None

    def load_embedding_model(self):
        """
        Load sentence transformer model for embeddings.

        Loads the all-MiniLM-L6-v2 model for semantic similarity calculations.
        Also pre-computes embeddings for executive summary detection terms.
        """
        if self.model is None:
            logger.info("Loading sentence transformer model (%s)...", EMBEDDING_MODEL)
            self.model = SentenceTransformer(EMBEDDING_MODEL)

            # Pre-compute embeddings for executive summary variations
            executive_summary_terms = [
                "executive summary",
                "report summary",
                "abstract",
                "overview",
                "key findings",
                "main findings",
                "conclusions",
            ]
            self.executive_summary_embeddings = self.model.encode(
                executive_summary_terms, convert_to_tensor=True
            )

            logger.info("✓ Model loaded")

    def detect_summary_sections(
        self, toc_text, language_code=None, threshold=0.65, page_count=None
    ):
        """
        Check if TOC contains summary sections using semantic similarity.

        Uses sentence transformer embeddings to identify sections like
        "Executive Summary", "Abstract", "Overview", etc. with semantic matching
        to handle variations in naming.

        Args:
            toc_text: Table of contents text
            language_code: ISO language code for the document (e.g., 'en', 'fr', 'es')
            threshold: Similarity threshold for matching (default: 0.65 = moderately similar)
            page_count: Optional page count for TOC validation

        Returns:
            tuple: (has_summary, section_headings)
                - has_summary: Boolean indicating if summary sections were detected
                - section_headings: Verbatim TOC lines for matched sections, or empty string
        """
        if not toc_text or not isinstance(toc_text, str):
            return False, ""

        # Count TOC entries (levels)
        lines = toc_text.split("\n")
        toc_entry_count = 0
        for line in lines:
            line = line.strip()
            if line and not line.startswith("---") and "[H" in line and "|" in line:
                toc_entry_count += 1

        # Validation: If TOC has < 5 entries and document has > 30 pages, something is wrong
        if page_count and toc_entry_count < 5 and page_count > 30:
            logger.warning(
                "  TOC validation failed: only %s entries for %s-page document"
                " - skipping TOC detection",
                toc_entry_count,
                page_count,
            )
            return False, ""

        logger.info("  TOC has %s entries", toc_entry_count)

        # Translate TOC to English first if not already English
        if language_code and language_code.lower() not in ["en", "eng", "english"]:
            try:
                # Strip heading markers ([H1], [H2], etc.) and page numbers before translation
                # Save all metadata to restore after translation
                lines_original = toc_text.split("\n")
                lines_text_only = []
                metadata = []  # Store (line_index, heading_marker, page_number) tuples

                for i, line in enumerate(lines_original):
                    # Parse format: "[H2] Heading text | page 5"
                    # Extract heading marker [HX]
                    heading_marker = None
                    heading_match = re.match(r"^(\s*)(\[H\d+\])\s*(.*)$", line)
                    if heading_match:
                        indent = heading_match.group(1)
                        heading_marker = heading_match.group(2)
                        rest_of_line = heading_match.group(3)
                    else:
                        indent = ""
                        rest_of_line = line

                    # Extract page number from end
                    page_num = None
                    page_match = re.match(
                        r"^(.*?)\s*\|\s*page\s+(\S+)\s*$", rest_of_line, re.IGNORECASE
                    )
                    if page_match:
                        text_only = page_match.group(1).strip()
                        page_num = page_match.group(2).strip()
                    else:
                        text_only = rest_of_line.strip()

                    # Store just the text for translation
                    lines_text_only.append(text_only)
                    # Store metadata to restore later
                    metadata.append((indent, heading_marker, page_num))

                # Translate only the heading text
                text_to_translate = "\n".join(lines_text_only)
                translated_text = self.translate_to_english(
                    text_to_translate, language_code
                )

                # Restore heading markers and page numbers to translated text
                translated_lines = translated_text.split("\n")
                toc_english_lines = []
                for i, translated_line in enumerate(translated_lines):
                    if i >= len(metadata):
                        # Safety: if translation added lines, just use them as-is
                        toc_english_lines.append(translated_line)
                        continue

                    indent, heading_marker, page_num = metadata[i]

                    # Rebuild the line with original format
                    reconstructed_line = ""
                    if indent:
                        reconstructed_line += indent
                    if heading_marker:
                        reconstructed_line += f"{heading_marker} "
                    reconstructed_line += translated_line.strip()
                    if page_num:
                        reconstructed_line += f" | page {page_num}"

                    toc_english_lines.append(reconstructed_line)

                toc_english = "\n".join(toc_english_lines)
                logger.info("  ✓ TOC translated successfully")

            except Exception as e:
                logger.warning("  Translation failed, using original TOC: %s", e)
                toc_english = toc_text
        else:
            toc_english = toc_text

        # Split TOC into lines and extract heading text
        lines = toc_english.split("\n")
        headings_with_lines = []  # Store tuples of (heading_text, original_line)
        for line in lines:
            line = line.strip()
            if not line or line.startswith("---"):
                continue
            # Extract heading text (remove level markers and page numbers)
            # Format: "[H1] Heading text | page 5"
            if "|" in line:
                heading_part = line.split("|")[0]
                # Remove [H1], [H2], etc.
                if "]" in heading_part:
                    heading_text = heading_part.split("]", 1)[1].strip()
                    headings_with_lines.append((heading_text.lower(), line))

        if not headings_with_lines:
            return False, ""

        # Check each heading individually against summary terms
        matched_lines = []
        for heading_text, original_line in headings_with_lines:
            # Encode the individual heading
            heading_embedding = self.model.encode(
                [heading_text], convert_to_tensor=True
            )

            # Compute similarity with summary terms
            similarities = util.cos_sim(
                heading_embedding, self.executive_summary_embeddings
            )

            # Check if this heading is similar to any summary term
            max_similarity = similarities.max().item()
            if max_similarity >= threshold:
                matched_lines.append(original_line)

        if matched_lines:
            return True, "\n".join(matched_lines)
        return False, ""

    def load_llm_model(self):
        """
        Load LLM model configuration and get HuggingFace token.

        Retrieves the HF_TOKEN or HUGGINGFACE_API_KEY from environment variables
        for accessing the HuggingFace Router API.

        Raises:
            ValueError: If no HuggingFace token is found in environment variables
        """
        # Get HuggingFace token from environment
        self.hf_token = os.getenv("HUGGINGFACE_API_KEY") or os.getenv("HF_TOKEN")
        if not self.hf_token:
            raise ValueError(
                "HUGGINGFACE_API_KEY or HF_TOKEN not found in environment. "
                "Please set it in .env file or environment variables. "
                "Get your token at https://huggingface.co/settings/tokens"
            )

        logger.info("✓ HuggingFace token loaded for model: %s", LLM_MODEL["model"])

    def translate_to_english(self, text, source_language):
        """
        Translate text to English if not already in English.

        Uses Google Translate via deep-translator library. Handles long texts
        by chunking into 5000-character segments.

        Args:
            text: Text to translate
            source_language: ISO language code (e.g., 'fr', 'es', 'ar')

        Returns:
            str: Translated text, or original if already English or translation fails
        """
        if not text or not source_language:
            return text

        # Normalize language code
        source_language = source_language.lower().strip()

        # If already English, no translation needed
        if source_language in ["en", "eng", "english"]:
            return text

        try:
            logger.info("  Translating from %s to English...", source_language)

            # Google Translate has character limits, so chunk if needed
            MAX_CHARS = 5000
            if len(text) <= MAX_CHARS:
                translator = GoogleTranslator(source=source_language, target="en")
                translated = translator.translate(text)
                logger.info("  ✓ Translation complete: %s chars", len(translated))
                return translated
            else:
                # Split into chunks
                chunks = []
                start = 0
                while start < len(text):
                    end = min(start + MAX_CHARS, len(text))
                    chunks.append(text[start:end])
                    start = end

                logger.info("  Translating in %s chunks...", len(chunks))
                translator = GoogleTranslator(source=source_language, target="en")
                translated_chunks = []
                for i, chunk in enumerate(chunks, 1):
                    translated = translator.translate(chunk)
                    translated_chunks.append(translated)
                    logger.info("  ✓ Chunk %s/%s translated", i, len(chunks))

                full_translation = " ".join(translated_chunks)
                logger.info("  ✓ Translation complete: %s chars", len(full_translation))
                return full_translation

        except Exception as e:
            logger.warning("  ✗ Translation failed: %s", str(e))
            logger.warning("  Using original text")
            return text

    def load_metadata(self):
        """
        Load metadata Excel file.

        Returns:
            tuple: (workbook, worksheet) where worksheet is the PDF Metadata sheet
        """
        logger.info("Loading metadata from: %s", self.metadata_path)
        wb = openpyxl.load_workbook(self.metadata_path)

        # Try to find the PDF Metadata sheet
        if "PDF Metadata" in wb.sheetnames:
            ws = wb["PDF Metadata"]
        elif "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
        else:
            ws = wb.active

        logger.info("✓ Loaded worksheet: %s", ws.title)
        return wb, ws

    def load_markdown(self, filepath):
        """
        Load and clean markdown content.

        Removes image references, HTML comments, and page separators from
        the markdown file.

        Args:
            filepath: Path to the markdown file

        Returns:
            str: Cleaned markdown content
        """
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

        # Remove image references
        content = re.sub(r"!\[.*?\]\(.*?\)", "", content)

        # Remove HTML comments
        content = re.sub(r"<!--.*?-->", "", content, flags=re.DOTALL)

        # Remove page separators (with or without page numbers)
        content = re.sub(r"------- Page \d+ -------", "", content)
        content = re.sub(r"------- Page Break -------", "", content)

        return content.strip()

    def tokenize_sentences(self, text):
        """
        Tokenize text into sentences.

        Uses NLTK's sentence tokenizer and filters out short sentences,
        tables, and separators.

        Args:
            text: Text to tokenize

        Returns:
            list: List of sentence strings
        """
        sentences = sent_tokenize(text)

        # Filter out unwanted sentences
        filtered = []
        for s in sentences:
            # Skip very short sentences
            if len(s.split()) <= 5:
                continue
            # Skip sentences with table markers or separators
            if "|" in s or "---" in s:
                continue
            filtered.append(s)

        return filtered

    def centroid_summary(self, content, num_sentences=None, doc_title=None):
        """
        Generate summary using centroid-based approach.

        Selects the most representative sentences by computing sentence embeddings
        and finding those closest to the document centroid (mean embedding).

        Args:
            content: Markdown document content
            num_sentences: Number of sentences to extract (default: NUM_CENTROID_SENTENCES)
            doc_title: Optional document title to prepend to summary

        Returns:
            str: Extractive summary with top sentences, or None if no valid sentences found
        """
        if num_sentences is None:
            num_sentences = NUM_CENTROID_SENTENCES

        sentences = self.tokenize_sentences(content)

        if len(sentences) == 0:
            return None

        logger.info("  Processing %s sentences...", len(sentences))

        # Encode sentences
        embeddings = self.model.encode(sentences, show_progress_bar=False)

        # Calculate centroid
        centroid = np.mean(embeddings, axis=0)

        # Calculate similarity to centroid
        similarities = util.cos_sim(centroid, embeddings)[0]

        # Get top sentences
        top_indices = similarities.argsort(descending=True)[:num_sentences]
        top_indices = sorted(top_indices)  # Keep original order

        # Clean up sentences and join
        summary_sentences = [sentences[i] for i in top_indices]
        # Remove extra whitespace and newlines within each sentence
        cleaned_sentences = [" ".join(s.split()) for s in summary_sentences]
        # Join with visual separator to add breaks between sentences
        summary = "\n\n---\n\n".join(cleaned_sentences)

        # Prepend document title if available
        if doc_title:
            summary = f"# {doc_title}\n\n{summary}"

        return summary

    def extract_summary_section(
        self, content, summary_headings, full_toc=None, doc_title=None
    ):
        """
        Extract content from identified summary sections.

        Parses the markdown content to find and extract sections matching
        the detected summary headings (Executive Summary, Abstract, etc.).
        Includes all subsections until the next major heading.

        Args:
            content: Full markdown content
            summary_headings: Newline-separated list of section headings to extract
                (from Key Content Sections column)
            full_toc: Full table of contents (optional, for future enhancements)
            doc_title: Optional document title to prepend to summary

        Returns:
            str: Extracted section content with headings, or None if no sections found
        """
        if (
            not summary_headings
            or summary_headings == "No"
            or summary_headings == "No TOC"
        ):
            return None

        # Parse summary headings (can be multiple, one per line)
        headings = summary_headings.strip().split("\n")

        # Extract heading text from TOC format: [H1] Executive summary | page 10
        target_headings = []
        for heading in headings:
            heading_match = re.search(r"\[H\d+\]\s*(.+?)\s*\|", heading)
            if heading_match:
                target_headings.append(heading_match.group(1).strip().lower())
            else:
                target_headings.append(heading.strip().lower())

        logger.info("  Target headings to extract: %s", target_headings)

        # Split content by all headings (any level)
        # Pattern to match markdown headings
        heading_pattern = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)

        # Find all headings with their positions
        heading_matches = list(heading_pattern.finditer(content))

        # Extract sections that match our target headings
        extracted_sections = []

        for i, match in enumerate(heading_matches):
            heading_level = match.group(1)  # Number of # symbols
            heading_text = match.group(2).strip().lower()

            # Check if this heading matches any of our target headings
            is_target = False
            matched_target = None
            for target in target_headings:
                if target in heading_text or heading_text in target:
                    is_target = True
                    matched_target = target
                    break

            if is_target:
                logger.info(
                    "  Matched heading '%s' (level %s) to target '%s' at position %s",
                    heading_text,
                    len(heading_level),
                    matched_target,
                    match.start(),
                )
                # Extract content from this heading to the next heading at same or higher level
                start_pos = match.end()
                end_pos = len(content)
                heading_depth = len(heading_level)

                # Find the next heading at same or higher level (this ends our section)
                for j in range(i + 1, len(heading_matches)):
                    next_match = heading_matches[j]
                    next_level = len(next_match.group(1))

                    # Stop at any heading of same or higher level (fewer or equal # symbols)
                    if next_level <= heading_depth:
                        end_pos = next_match.start()
                        logger.info(
                            "    Section ends at heading '%s' (level %s) at pos %s",
                            next_match.group(2).strip(),
                            next_level,
                            end_pos,
                        )
                        break

                if end_pos == len(content):
                    logger.info(
                        "    Section extends to end of document (pos %s)", end_pos
                    )

                # Extract the section content
                section_content = content[start_pos:end_pos].strip()

                logger.info(
                    "  Extracted section '%s' from pos %s to %s (%s chars)",
                    heading_text,
                    start_pos,
                    end_pos,
                    len(section_content),
                )

                if section_content:
                    # Add heading and content
                    extracted_sections.append(
                        f"{match.group(1)} {match.group(2)}\n\n{section_content}"
                    )
                else:
                    logger.warning("  ⚠ Section '%s' has no content!", heading_text)

        if not extracted_sections:
            logger.warning(
                "  Could not extract summary sections using heading matching"
            )
            return None

        # Join all extracted sections
        full_text = "\n\n".join(extracted_sections)

        logger.info(
            "  ✓ Extracted %s section(s), total %s chars",
            len(extracted_sections),
            len(full_text),
        )

        # Clean up any remaining page breaks that might have been in the sections
        full_text = re.sub(r"------- Page \d+ -------", "", full_text)
        full_text = re.sub(r"------- Page Break -------", "", full_text)

        # Remove multiple consecutive line breaks (more than 2 newlines)
        full_text = re.sub(r"\n\s*\n\s*\n+", "\n\n", full_text)

        # Prepend document title if available
        if doc_title:
            full_text = f"# {doc_title}\n\n{full_text.strip()}"

        return full_text.strip()

    def llm_summary_chat(self, content):
        """
        Generate summary using HuggingFace Router API (Mistral/Llama).

        Uses chat-based LLM models for abstractive summarization. Automatically
        handles documents that exceed the model's context window by implementing
        a map-reduce strategy (chunking + merging).

        Args:
            content: Text from 'Key Content (using summary sections)' column

        Returns:
            tuple: (final_summary, chunked_summary)
                - final_summary: Final abstractive summary
                - chunked_summary: Combined chunk summaries (None for single-pass)
                - Returns (None, None) if content is empty
                - Returns ("USE_CENTROID", None) if content requires too many chunks

        Raises:
            RuntimeError: If LLM API call fails (will abort summarization)
        """
        if not content or content in ["No", "No TOC"]:
            logger.info("  LLM summary skipped: no content or TOC")
            return None, None

        # Clean content
        logger.info("  Cleaning content...")
        cleaned = re.sub(r"!\[.*?\]\(.*?\)", "", content)  # Remove images
        cleaned = re.sub(r"<!--.*?-->", "", cleaned, flags=re.DOTALL)  # Remove comments
        cleaned = re.sub(r"------- Page \d+ -------", "", cleaned)  # Remove page breaks
        cleaned = re.sub(
            r"```.*?```", "", cleaned, flags=re.DOTALL
        )  # Remove code blocks
        cleaned = cleaned.strip()

        # Remove excessive whitespace
        cleaned = re.sub(r"\n\s*\n\s*\n+", "\n\n", cleaned)
        cleaned = re.sub(r" +", " ", cleaned)
        cleaned = cleaned.strip()

        if not cleaned:
            logger.warning("  Content empty after cleaning")
            return None, None

        logger.info("  Input: %s characters", len(cleaned))

        # Get max context chars from config
        MAX_CHARS = LLM_MODEL["max_context_chars"]

        # For single-pass, we use reduction prompt with the full text
        # Account for prompt overhead
        prompt_overhead = len(LLM_REDUCTION_PROMPT) + 100
        EFFECTIVE_MAX_CHARS = MAX_CHARS - prompt_overhead

        logger.info(
            "  Max context: %s chars, effective after prompt overhead: %s chars",
            MAX_CHARS,
            EFFECTIVE_MAX_CHARS,
        )

        # API endpoint and headers
        API_URL = "https://router.huggingface.co/v1/chat/completions"
        headers = {"Authorization": f"Bearer {self.hf_token}"}

        try:
            if len(cleaned) <= EFFECTIVE_MAX_CHARS:
                # Input fits in one pass
                logger.info(
                    "  Processing in single pass (input fits within %s char limit)",
                    EFFECTIVE_MAX_CHARS,
                )

                # Format the reduction prompt with the document text
                formatted_prompt = LLM_REDUCTION_PROMPT.replace(
                    "{document_text}", cleaned
                )

                # Make API request
                payload = {
                    "model": LLM_MODEL["model"],
                    "messages": [
                        {"role": "user", "content": formatted_prompt},
                    ],
                    "max_tokens": LLM_MODEL["max_tokens"],
                    "temperature": LLM_TEMPERATURE,
                }

                response = requests.post(
                    API_URL, headers=headers, json=payload, timeout=120
                )
                response.raise_for_status()
                result = response.json()

                if "error" in result:
                    raise ValueError(f"API error: {result['error']}")

                summary = result["choices"][0]["message"]["content"].strip()

                # Clean up markdown formatting
                summary = clean_markdown_formatting(summary)

                if not summary or len(summary) < 50:
                    raise ValueError(f"LLM response too short: {len(summary)} chars")

                logger.info("  ✓ Generated summary: %s characters", len(summary))
                logger.info("  Full summary text:\n%s", summary)
                return summary, None  # No chunked summary for single-pass
            else:
                # Input too long - use map-reduce with chunking
                logger.info(
                    "  Input too long (%s chars), using map-reduce strategy",
                    len(cleaned),
                )

                # Split into chunks
                chunks = []
                start = 0
                overlap = LLM_MODEL["chunk_overlap"]

                while start < len(cleaned):
                    end = min(start + EFFECTIVE_MAX_CHARS, len(cleaned))
                    chunks.append(cleaned[start:end])
                    if end >= len(cleaned):
                        break
                    start = end - overlap

                logger.info(
                    "  Split into %s chunks (overlap=%s chars)", len(chunks), overlap
                )

                # Check if too many chunks
                if len(chunks) > 30:
                    logger.warning(
                        "  ✗ Content requires %s chunks (max 30) - will use centroid fallback",
                        len(chunks),
                    )
                    return "USE_CENTROID", None

                # MAP: Summarize each chunk
                logger.info("  MAP phase: Summarizing each chunk...")
                chunk_summaries = []
                for i, chunk in enumerate(chunks, 1):
                    logger.info(
                        "    Processing chunk %s/%s (%s chars)...",
                        i,
                        len(chunks),
                        len(chunk),
                    )

                    # Format the reduction prompt with chunk text
                    formatted_prompt = LLM_REDUCTION_PROMPT.replace(
                        "{document_text}", chunk
                    )

                    payload = {
                        "model": LLM_MODEL["model"],
                        "messages": [
                            {"role": "user", "content": formatted_prompt},
                        ],
                        "max_tokens": LLM_MODEL.get("chunk_tokens", 800),
                        "temperature": LLM_TEMPERATURE,
                    }

                    response = requests.post(
                        API_URL, headers=headers, json=payload, timeout=120
                    )
                    response.raise_for_status()
                    result = response.json()

                    if "error" in result:
                        raise ValueError(f"API error on chunk {i}: {result['error']}")

                    chunk_summary = result["choices"][0]["message"]["content"].strip()

                    if not chunk_summary or len(chunk_summary) < 20:
                        raise ValueError(
                            f"LLM chunk {i} response too short: {len(chunk_summary)} chars"
                        )

                    chunk_summaries.append(chunk_summary)
                    logger.info(
                        "    ✓ Chunk %s summary: %s chars", i, len(chunk_summary)
                    )

                # REDUCE: Combine chunk summaries
                logger.info(
                    "  REDUCE phase: Combining %s chunk summaries...",
                    len(chunk_summaries),
                )
                combined = "\n\n".join(chunk_summaries)

                # Final summarization if combined is still manageable
                if len(combined) <= MAX_CHARS:
                    # Format the summary prompt with the map summaries
                    formatted_prompt = LLM_SUMMARY_PROMPT.replace(
                        "{map_summaries}", combined
                    )

                    payload = {
                        "model": LLM_MODEL["model"],
                        "messages": [
                            {"role": "user", "content": formatted_prompt},
                        ],
                        "max_tokens": LLM_MODEL["max_tokens"],
                        "temperature": LLM_TEMPERATURE,
                    }

                    response = requests.post(
                        API_URL, headers=headers, json=payload, timeout=120
                    )
                    response.raise_for_status()
                    result = response.json()

                    if "error" in result:
                        raise ValueError(
                            f"API error on final summary: {result['error']}"
                        )

                    final_summary = result["choices"][0]["message"]["content"].strip()

                    # Clean up markdown formatting
                    final_summary = clean_markdown_formatting(final_summary)

                    if not final_summary or len(final_summary) < 50:
                        raise ValueError(
                            f"LLM final response too short: {len(final_summary)} chars"
                        )

                    logger.info("  ✓ Final summary: %s characters", len(final_summary))
                    logger.info("  Full summary text:\n%s", final_summary)
                    return final_summary, combined
                else:
                    logger.warning(
                        "  Combined summaries too long, returning chunk summaries"
                    )
                    return combined, combined

        except Exception as e:
            logger.error("  ✗ CRITICAL: LLM summarization failed: %s", e)
            logger.error("  Aborting summarization due to LLM failure")
            import traceback

            traceback.print_exc()
            # Re-raise the exception to abort the entire process
            raise RuntimeError(f"LLM API call failed: {e}") from e

    def llm_summary(self, content):
        """
        Generate summary using LLM (routes to appropriate model type).

        Main entry point for LLM summarization. Routes to the appropriate
        implementation based on the model type (BART vs chat-based models).

        Args:
            content: Text from 'Key Content (using summary sections)' column

        Returns:
            tuple: (final_summary, chunked_summary) or (final_summary, None) for single-pass
                Returns (None, None) if no content
        """
        # Route to appropriate method based on model type
        if self.model_type == "bart":
            return self.llm_summary_bart(content)
        else:
            # Mistral, Llama, and other chat models
            return self.llm_summary_chat(content)

    def llm_summary_bart(self, content):
        """
        Generate summary using HuggingFace BART model via Inference API.

        Uses BART specialized summarization model. Implements map-reduce strategy
        for documents exceeding the 4K character context window.

        Args:
            content: Text from 'Key Content (using summary sections)' column

        Returns:
            tuple: (final_summary, chunked_summary) or (final_summary, None) for single-pass
                Returns (None, None) if no content
                Returns ("USE_CENTROID", None) if content requires too many chunks

        Raises:
            RuntimeError: If LLM API call fails (will abort summarization)
        """
        if not content or content in ["No", "No TOC"]:
            logger.info("  LLM summary skipped: no content or TOC")
            return None, None

        # Clean content
        logger.info("  Cleaning content...")
        cleaned = re.sub(r"!\[.*?\]\(.*?\)", "", content)  # Remove images
        cleaned = re.sub(r"<!--.*?-->", "", cleaned, flags=re.DOTALL)  # Remove comments
        cleaned = re.sub(r"------- Page \d+ -------", "", cleaned)  # Remove page breaks
        cleaned = re.sub(
            r"```.*?```", "", cleaned, flags=re.DOTALL
        )  # Remove code blocks
        cleaned = cleaned.strip()

        # Remove excessive whitespace and blank lines
        cleaned = re.sub(r"\n\s*\n\s*\n+", "\n\n", cleaned)
        cleaned = re.sub(r" +", " ", cleaned)
        cleaned = cleaned.strip()

        if not cleaned:
            logger.warning("  Content empty after cleaning")
            return None, None

        logger.info("  Input: %s characters", len(cleaned))

        try:
            # BART can handle ~1024 tokens (~4000 characters)
            MAX_CHARS = LLM_MODEL["max_context_chars"]

            if len(cleaned) <= MAX_CHARS:
                # Input fits in one pass
                logger.info(
                    "  Processing in single pass (input fits within %s char limit)",
                    MAX_CHARS,
                )

                # API endpoint and headers for BART
                API_URL = f"https://router.huggingface.co/hf-inference/models/{LLM_MODEL['model']}"
                headers = {"Authorization": f"Bearer {self.hf_token}"}

                payload = {"inputs": cleaned}

                response = requests.post(
                    API_URL, headers=headers, json=payload, timeout=120
                )

                if response.status_code != 200:
                    logger.error("  API Error Response: %s", response.text)

                response.raise_for_status()
                result = response.json()

                if isinstance(result, dict) and "error" in result:
                    raise ValueError(f"API error: {result['error']}")

                # Extract summary text
                if isinstance(result, list) and len(result) > 0:
                    summary = result[0].get("summary_text", str(result[0]))
                elif isinstance(result, dict):
                    summary = result.get("summary_text", str(result))
                else:
                    summary = str(result)

                summary = summary.strip()

                # Clean up markdown formatting
                summary = clean_markdown_formatting(summary)

                if not summary or len(summary) < 50:
                    raise ValueError(f"BART response too short: {len(summary)} chars")

                logger.info("  ✓ Generated summary: %s characters", len(summary))
                logger.info("  Full summary text:\n%s", summary)
                return summary, None  # No chunked summary for single-pass
            else:
                # Input too long - use map-reduce with chunking
                logger.info(
                    "  Input too long (%s chars), using map-reduce strategy",
                    len(cleaned),
                )

                # Split into chunks by characters
                chunks = []
                start = 0
                overlap = LLM_MODEL["chunk_overlap"]

                while start < len(cleaned):
                    end = min(start + MAX_CHARS, len(cleaned))
                    chunks.append(cleaned[start:end])
                    if end >= len(cleaned):
                        break
                    start = end - overlap

                logger.info(
                    "  Split into %s chunks (overlap=%s chars)", len(chunks), overlap
                )
                for i, chunk in enumerate(chunks, 1):
                    logger.info("    Chunk %s: %s chars", i, len(chunk))

                # Check if too many chunks
                if len(chunks) > 30:
                    logger.warning(
                        "  ✗ Content requires %s chunks (max 30) - "
                        "will use centroid fallback",
                        len(chunks),
                    )
                    return "USE_CENTROID", None

                # MAP: Summarize each chunk
                logger.info("  MAP phase: Summarizing each chunk...")
                chunk_summaries = []
                for i, chunk in enumerate(chunks, 1):
                    logger.info(
                        "    Processing chunk %s/%s (%s chars)...",
                        i,
                        len(chunks),
                        len(chunk),
                    )

                    # API endpoint and headers for BART
                    API_URL = (
                        f"https://router.huggingface.co/hf-inference/"
                        f"models/{LLM_MODEL['model']}"
                    )
                    headers = {"Authorization": f"Bearer {self.hf_token}"}

                    payload = {"inputs": chunk}

                    response = requests.post(
                        API_URL, headers=headers, json=payload, timeout=120
                    )

                    if response.status_code != 200:
                        logger.error("  API Error Response: %s", response.text)

                    response.raise_for_status()
                    result = response.json()

                    if isinstance(result, dict) and "error" in result:
                        raise ValueError(f"API error on chunk {i}: {result['error']}")

                    # Extract summary
                    if isinstance(result, list) and len(result) > 0:
                        chunk_summary = result[0].get("summary_text", str(result[0]))
                    elif isinstance(result, dict):
                        chunk_summary = result.get("summary_text", str(result))
                    else:
                        chunk_summary = str(result)

                    chunk_summary = chunk_summary.strip()

                    if not chunk_summary or len(chunk_summary) < 20:
                        raise ValueError(
                            f"BART chunk {i} response too short: {len(chunk_summary)} chars"
                        )

                    chunk_summaries.append(chunk_summary)
                    logger.info(
                        "    ✓ Chunk %s summary: %s chars", i, len(chunk_summary)
                    )

                # REDUCE: Combine chunk summaries
                logger.info(
                    "  REDUCE phase: Combining %s chunk summaries...",
                    len(chunk_summaries),
                )
                combined = "\n\n".join(chunk_summaries)
                logger.info("  Combined length: %s characters", len(combined))

                # Check if combined text is still too long
                if len(combined) > MAX_CHARS:
                    logger.info(
                        "  Combined text still too long (%s chars), "
                        "applying second round of map-reduce",
                        len(combined),
                    )

                    # Split combined text into chunks again
                    reduce_chunks = []
                    start = 0
                    while start < len(combined):
                        end = min(start + MAX_CHARS, len(combined))
                        reduce_chunks.append(combined[start:end])
                        if end >= len(combined):
                            break
                        start = end - overlap

                    logger.info(
                        "  Split combined text into %s chunks for second reduction",
                        len(reduce_chunks),
                    )

                    # Summarize each chunk of the combined text
                    reduce_summaries = []
                    for i, chunk in enumerate(reduce_chunks, 1):
                        logger.info(
                            "    Reducing chunk %s/%s (%s chars)...",
                            i,
                            len(reduce_chunks),
                            len(chunk),
                        )

                        # API endpoint and headers for BART
                        model_name = LLM_MODEL["model"]
                        API_URL = f"https://router.huggingface.co/hf-inference/models/{model_name}"
                        headers = {"Authorization": f"Bearer {self.hf_token}"}

                        payload = {"inputs": chunk}

                        response = requests.post(
                            API_URL, headers=headers, json=payload, timeout=120
                        )
                        response.raise_for_status()
                        result = response.json()

                        if isinstance(result, dict) and "error" in result:
                            raise ValueError(
                                f"API error on round 2 chunk {i}: {result['error']}"
                            )

                        if isinstance(result, list) and len(result) > 0:
                            reduce_summary = result[0].get(
                                "summary_text", str(result[0])
                            )
                        elif isinstance(result, dict):
                            reduce_summary = result.get("summary_text", str(result))
                        else:
                            reduce_summary = str(result)

                        reduce_summaries.append(reduce_summary.strip())
                        logger.info(
                            "    ✓ Reduce chunk %s summary: %s chars",
                            i,
                            len(reduce_summary),
                        )

                    # Final reduction
                    final_combined = "\n\n".join(reduce_summaries)
                    logger.info("  Final combined text: %s chars", len(final_combined))

                    # Final summarization pass
                    logger.info(
                        "  Final summarization pass (after second reduction)..."
                    )

                    # API endpoint and headers for BART
                    API_URL = (
                        f"https://router.huggingface.co/hf-inference/"
                        f"models/{LLM_MODEL['model']}"
                    )
                    headers = {"Authorization": f"Bearer {self.hf_token}"}

                    payload = {"inputs": final_combined}

                    response = requests.post(
                        API_URL, headers=headers, json=payload, timeout=120
                    )
                    response.raise_for_status()
                    result = response.json()

                    if isinstance(result, dict) and "error" in result:
                        raise ValueError(
                            f"API error on final summary: {result['error']}"
                        )

                    if isinstance(result, list) and len(result) > 0:
                        final_summary = result[0].get("summary_text", str(result[0]))
                    elif isinstance(result, dict):
                        final_summary = result.get("summary_text", str(result))
                    else:
                        final_summary = str(result)

                    final_summary = final_summary.strip()

                    # Clean up markdown formatting
                    final_summary = clean_markdown_formatting(final_summary)

                    logger.info(
                        "  ✓ Generated final summary: %s characters", len(final_summary)
                    )
                    logger.info("  Full final summary:\n%s", final_summary)
                    return final_summary, combined
                else:
                    # Combined text fits - do final summarization
                    logger.info("  Final summarization pass...")

                    # API endpoint and headers for BART
                    API_URL = (
                        f"https://router.huggingface.co/hf-inference/"
                        f"models/{LLM_MODEL['model']}"
                    )
                    headers = {"Authorization": f"Bearer {self.hf_token}"}

                    payload = {"inputs": combined}

                    response = requests.post(
                        API_URL, headers=headers, json=payload, timeout=120
                    )
                    response.raise_for_status()
                    result = response.json()

                    if isinstance(result, dict) and "error" in result:
                        raise ValueError(
                            f"API error on final summary: {result['error']}"
                        )

                    if isinstance(result, list) and len(result) > 0:
                        final_summary = result[0].get("summary_text", str(result[0]))
                    elif isinstance(result, dict):
                        final_summary = result.get("summary_text", str(result))
                    else:
                        final_summary = str(result)

                    final_summary = final_summary.strip()

                    # Clean up markdown formatting
                    final_summary = clean_markdown_formatting(final_summary)

                    logger.info(
                        "  ✓ Generated final summary: %s characters", len(final_summary)
                    )
                    logger.info("  Full final summary:\n%s", final_summary)
                    return final_summary, combined

        except Exception as e:
            logger.error("  ✗ CRITICAL: LLM (BART) summarization failed: %s", e)
            logger.error("  Aborting summarization due to LLM failure")
            import traceback

            logger.error(traceback.format_exc())
            # Re-raise the exception to abort the entire process
            raise RuntimeError(f"LLM API call failed: {e}") from e

    def ensure_columns_exist(self, ws, wb):
        """
        Ensure summary columns exist in the worksheet.

        Adds any missing summary columns to the Excel metadata sheet header row.

        Args:
            ws: Excel worksheet
            wb: Excel workbook
        """
        headers = [cell.value for cell in ws[1]]
        logger.info("Current headers before ensuring columns: %s", headers)

        required_columns = [
            "Key Content Sections",
            "Extractive Summary (using summary TOC sections)",
            "Extractive Summary (using centroid distance)",
            "Abstractive Summary Input Method",
            "Abstractive Summary (concatenating chunk summaries)",
            "Abstractive Summary (map reduced)",
        ]

        # Find the last non-empty column
        last_col_idx = len(headers)
        # Trim trailing None/empty values from headers
        while last_col_idx > 0 and (
            headers[last_col_idx - 1] is None or headers[last_col_idx - 1] == ""
        ):
            last_col_idx -= 1

        # Rebuild headers list without trailing empties
        headers = headers[:last_col_idx]

        columns_added = False
        for column_name in required_columns:
            if column_name not in headers:
                # Add at the end of actual columns
                col_idx = len(headers) + 1
                ws.cell(1, col_idx, column_name)
                headers.append(column_name)
                logger.info("Added '%s' column at index %s", column_name, col_idx)
                columns_added = True

        if columns_added:
            wb.save(self.metadata_path)
            logger.info("✓ Columns added and saved")
        else:
            logger.info("All required columns already exist")

        logger.info(
            "Final headers after ensuring columns: %s",
            [cell.value for cell in ws[1] if cell.value],
        )

    def summarize_all(self, force_summarize=False):
        """
        Summarize all documents in the metadata sheet.

        Main entry point for batch summarization. Processes each document in the
        metadata Excel file, generating:
        - Key content sections detection
        - Extractive summaries (centroid and section-based)
        - Abstractive summaries (LLM-generated)
        - Translation to English for non-English documents

        Args:
            force_summarize: If True, re-summarize documents even if already processed
        """
        # Load models
        self.load_embedding_model()
        self.load_llm_model()

        # Load metadata
        wb, ws = self.load_metadata()

        # Ensure columns exist
        self.ensure_columns_exist(ws, wb)

        # Get column indices (AFTER ensuring columns exist)
        headers = [cell.value for cell in ws[1]]

        if "Parsed Markdown Path" in headers:
            markdown_path_col = headers.index("Parsed Markdown Path") + 1
        elif "parsed_markdown_path" in headers:
            markdown_path_col = headers.index("parsed_markdown_path") + 1
        else:
            logger.error("Required column 'Parsed Markdown Path' not found")
            return

        language_col = None
        if "Language" in headers:
            language_col = headers.index("Language") + 1

        title_col = None
        if "Title" in headers:
            title_col = headers.index("Title") + 1

        toc_col = None
        if "TOC" in headers:
            toc_col = headers.index("TOC") + 1

        key_sections_col = headers.index("Key Content Sections") + 1

        # Handle both old and new column names for backwards compatibility
        if "Extractive Summary (using summary TOC sections)" in headers:
            summary_section_col = (
                headers.index("Extractive Summary (using summary TOC sections)") + 1
            )
        elif "Key Content (using key content sections)" in headers:
            summary_section_col = (
                headers.index("Key Content (using key content sections)") + 1
            )
        elif "Key Content (using summary sections)" in headers:
            summary_section_col = (
                headers.index("Key Content (using summary sections)") + 1
            )
        else:
            logger.error("Could not find summary section content column")
            return

        # Handle both old and new column names for backwards compatibility
        if "Extractive Summary (using centroid distance)" in headers:
            centroid_col = (
                headers.index("Extractive Summary (using centroid distance)") + 1
            )
        elif "Key Content (centroid all)" in headers:
            centroid_col = headers.index("Key Content (centroid all)") + 1
        else:
            logger.error("Could not find centroid column")
            return

        # Handle both old and new column names for backwards compatibility
        if "Abstractive Summary Input Method" in headers:
            llm_input_method_col = headers.index("Abstractive Summary Input Method") + 1
        elif "LLM input method" in headers:
            llm_input_method_col = headers.index("LLM input method") + 1
        else:
            logger.error("Could not find LLM input method column")
            return

        # Handle both old and new column names for backwards compatibility
        if "Abstractive Summary (concatenating chunk summaries)" in headers:
            llm_chunked_col = (
                headers.index("Abstractive Summary (concatenating chunk summaries)") + 1
            )
        elif "Summary (combining LLM summaries of chunks)" in headers:
            llm_chunked_col = (
                headers.index("Summary (combining LLM summaries of chunks)") + 1
            )
        elif "Summary (LLM chunked)" in headers:
            llm_chunked_col = headers.index("Summary (LLM chunked)") + 1
        else:
            logger.error("Could not find LLM chunked column")
            return

        # Handle both old and new column names for backwards compatibility
        if "Abstractive Summary (map reduced)" in headers:
            llm_combined_col = headers.index("Abstractive Summary (map reduced)") + 1
        elif "Summary (LLM map-reduced)" in headers:
            llm_combined_col = headers.index("Summary (LLM map-reduced)") + 1
        elif "Summary (LLM combined)" in headers:
            llm_combined_col = headers.index("Summary (LLM combined)") + 1
        else:
            logger.error("Could not find LLM combined column")
            return

        logger.info(
            "Column indices - Centroid: %s, Summary Section: %s, "
            "LLM Chunked: %s, LLM Combined: %s",
            centroid_col,
            summary_section_col,
            llm_chunked_col,
            llm_combined_col,
        )

        # Process each row
        total_rows = ws.max_row
        processed = 0
        skipped = 0

        for row_idx in range(2, total_rows + 1):
            markdown_path = ws.cell(row_idx, markdown_path_col).value

            if not markdown_path:
                skipped += 1
                continue

            # Check if already summarized (unless force)
            if not force_summarize:
                existing_centroid = ws.cell(row_idx, centroid_col).value
                if existing_centroid:
                    skipped += 1
                    continue

            # Load markdown
            logger.info(
                "\n[%s/%s] Processing: %s", row_idx - 1, total_rows - 1, markdown_path
            )

            content = self.load_markdown(markdown_path)
            if not content:
                logger.warning("  ✗ Could not load markdown")
                continue

            logger.info("  Loaded %s characters", len(content))

            # Get document title if available
            doc_title = None
            if title_col:
                doc_title = ws.cell(row_idx, title_col).value
                if doc_title:
                    logger.info("  Document title: %s", doc_title)

            # Count pages from content (look for "------- Page X -------" markers)
            page_count = len(re.findall(r"------- Page \d+ -------", content))
            if page_count > 0:
                logger.info("  Document has %s pages", page_count)

            # Detect summary sections from TOC
            if toc_col:
                toc_text = ws.cell(row_idx, toc_col).value
                if toc_text:
                    doc_language = None
                    if language_col:
                        doc_language = ws.cell(row_idx, language_col).value

                    logger.info("  Detecting summary sections from TOC...")
                    has_summary, section_headings = self.detect_summary_sections(
                        toc_text, doc_language, page_count=page_count
                    )

                    if has_summary:
                        ws.cell(row_idx, key_sections_col, section_headings)
                        logger.info(
                            "  ✓ Summary sections detected:\n%s", section_headings
                        )

                        # Save detected summary section headings to file for debugging
                        sections_toc_path = self.save_content_to_file(
                            markdown_path, section_headings, "summary_sections_toc"
                        )
                        if sections_toc_path:
                            logger.info(
                                "  ✓ Summary sections TOC saved to: %s",
                                sections_toc_path,
                            )
                    else:
                        ws.cell(row_idx, key_sections_col, "No")
                        logger.info("  No summary sections detected")
                else:
                    ws.cell(row_idx, key_sections_col, "No TOC")
                    logger.info("  No TOC available")
            else:
                logger.warning("  TOC column not found - skipping section detection")

            # Generate summaries
            logger.info("  Generating centroid summary...")
            centroid_summary = self.centroid_summary(content, doc_title=doc_title)
            if centroid_summary:
                # Save centroid summary to file instead of Excel cell
                centroid_path = self.save_content_to_file(
                    markdown_path, centroid_summary, "centroid"
                )
                if centroid_path:
                    logger.info(
                        "  Writing centroid path to row %s, col %s",
                        row_idx,
                        centroid_col,
                    )
                    ws.cell(row_idx, centroid_col, centroid_path)
                    logger.info(
                        "  ✓ Centroid summary saved to file: %s (%s chars)",
                        centroid_path,
                        len(centroid_summary),
                    )
                else:
                    logger.warning("  ✗ Failed to save centroid summary to file")
            else:
                logger.warning("  ✗ Centroid summary returned None")

            # Extract content from detected summary sections
            section_headings = ws.cell(row_idx, key_sections_col).value
            logger.info("  Summary section headings value: %s", section_headings)
            if section_headings and section_headings not in ["No", "No TOC"]:
                logger.info("  Generating summary section-based...")
                # Pass both the key sections AND the full TOC for context
                full_toc = ws.cell(row_idx, toc_col).value if toc_col else None
                section_summary = self.extract_summary_section(
                    content, section_headings, full_toc, doc_title=doc_title
                )
                if section_summary:
                    # Save section summary to file instead of Excel cell
                    section_path = self.save_content_to_file(
                        markdown_path, section_summary, "key_content_sections"
                    )
                    if section_path:
                        logger.info(
                            "  Writing section summary path to row %s, col %s",
                            row_idx,
                            summary_section_col,
                        )
                        ws.cell(row_idx, summary_section_col, section_path)
                        logger.info(
                            "  ✓ Section summary saved to file: %s (%s chars)",
                            section_path,
                            len(section_summary),
                        )
                    else:
                        logger.warning("  ✗ Failed to save section summary to file")
                else:
                    logger.warning("  ✗ Section summary returned None")
            else:
                logger.info(
                    "  Skipping section summary - headings are '%s'", section_headings
                )

            # Generate LLM summary (always enabled)
            logger.info("  Generating LLM summary...")
            # Use extractive summary based on EXTRACTIVE_SUMMARY_METHOD setting
            summary_section_path = ws.cell(row_idx, summary_section_col).value
            centroid_path = ws.cell(row_idx, centroid_col).value

            llm_input_method = None
            llm_input = None

            if EXTRACTIVE_SUMMARY_METHOD == "summary_sections":
                # Priority 1: Try to load from section summary file
                if summary_section_path and summary_section_path not in [
                    "No",
                    "No TOC",
                    None,
                    "",
                ]:
                    summary_section_content = self.load_content_from_file(
                        summary_section_path
                    )
                    if summary_section_content:
                        llm_input = summary_section_content
                        llm_input_method = "Using TOC summary sections content"
                        logger.info(
                            "  Using summary section content for LLM (%s chars)",
                            len(llm_input),
                        )
                    else:
                        # File path exists but couldn't load, fall through to centroid
                        logger.warning(
                            "  Could not load summary section content from: %s",
                            summary_section_path,
                        )

                # Priority 2: Fallback to centroid summary if no section content
                if not llm_input:
                    if centroid_path and centroid_path not in [None, ""]:
                        centroid_content = self.load_content_from_file(centroid_path)
                        if centroid_content:
                            llm_input = centroid_content
                            llm_input_method = "Using Centroid Content"
                            logger.info(
                                "  Using centroid content for LLM (%s chars)",
                                len(llm_input),
                            )
                        else:
                            # File path exists but couldn't load, fall through to full content
                            logger.warning(
                                "  Could not load centroid content from: %s",
                                centroid_path,
                            )

            elif EXTRACTIVE_SUMMARY_METHOD == "centroid":
                # Priority 1: Use centroid summary
                if centroid_path and centroid_path not in [None, ""]:
                    centroid_content = self.load_content_from_file(centroid_path)
                    if centroid_content:
                        llm_input = centroid_content
                        llm_input_method = "Using Centroid Content"
                        logger.info(
                            "  Using centroid content for LLM (%s chars)",
                            len(llm_input),
                        )
                    else:
                        # File path exists but couldn't load, fall through to full content
                        logger.warning(
                            "  Could not load centroid content from: %s", centroid_path
                        )

            # Final fallback to full content for both methods
            if not llm_input:
                llm_input = content
                llm_input_method = "Using Full Content"
                logger.info("  Using full content for LLM (%s chars)", len(llm_input))

            llm_summary, llm_chunked = self.llm_summary(llm_input)

            # Check if we need to use centroid fallback (content too large)
            if llm_summary == "USE_CENTROID":
                logger.info("  Content too large, falling back to centroid...")
                centroid_path = ws.cell(row_idx, centroid_col).value
                if centroid_path and centroid_path not in [None, ""]:
                    centroid_content = self.load_content_from_file(centroid_path)
                    if centroid_content:
                        llm_input_method = (
                            "Using Centroid Content (fallback from oversized)"
                        )
                        logger.info(
                            "  Retrying with centroid content (%s chars)",
                            len(centroid_content),
                        )
                        llm_summary, llm_chunked = self.llm_summary(centroid_content)
                    else:
                        logger.warning("  Could not load centroid for fallback")
                        llm_summary = None
                else:
                    logger.warning("  No centroid available for fallback")
                    llm_summary = None

            if llm_summary:
                # Get document language for translation
                doc_language = None
                if language_col:
                    doc_language = ws.cell(row_idx, language_col).value
                    if doc_language and doc_language != "Unknown":
                        logger.info("  Document language: %s", doc_language)

                # Translate summary to English if needed
                if doc_language and doc_language.lower() not in [
                    "en",
                    "eng",
                    "english",
                ]:
                    llm_summary = self.translate_to_english(llm_summary, doc_language)
                    if llm_chunked:
                        llm_chunked = self.translate_to_english(
                            llm_chunked, doc_language
                        )

                # Save chunked summary to file if it exists (from map-reduce)
                # For single-pass, llm_chunked will be None - use final summary as chunked too
                if llm_chunked:
                    # Save to file for backup/debugging (both doc folder and centralized)
                    llm_chunked_path, centralized_chunked = (
                        self.save_llm_summary_centralized(
                            markdown_path, llm_chunked, "llm_chunked_summary"
                        )
                    )
                    if llm_chunked_path:
                        logger.info(
                            "  ✓ LLM chunked summary saved to file: %s",
                            llm_chunked_path,
                        )

                    # Write actual summary content to Excel (not path)
                    logger.info(
                        "  Writing LLM chunked summary to row %s, col %s",
                        row_idx,
                        llm_chunked_col,
                    )
                    ws.cell(row_idx, llm_chunked_col, llm_chunked)
                    logger.info("  ✓ LLM chunked summary: %s chars", len(llm_chunked))
                else:
                    # Single-pass (no chunking) - use final summary for both columns
                    logger.info("  Single-pass summarization (no chunking needed)")
                    ws.cell(row_idx, llm_chunked_col, llm_summary)
                    logger.info(
                        "  ✓ Using final summary as chunked summary: %s chars",
                        len(llm_summary),
                    )

                # Save final summary to file (both doc folder and centralized)
                llm_summary_path, centralized_summary = (
                    self.save_llm_summary_centralized(
                        markdown_path, llm_summary, "llm_summary"
                    )
                )
                if llm_summary_path:
                    logger.info("  ✓ LLM summary saved to file: %s", llm_summary_path)

                # Write actual summary content to Excel (not path)
                logger.info(
                    "  Writing LLM summary to row %s, col %s", row_idx, llm_combined_col
                )
                ws.cell(row_idx, llm_combined_col, llm_summary)
                logger.info("  ✓ LLM summary: %s chars", len(llm_summary))

                # Save the input method used
                logger.info(
                    "  Writing LLM input method to row %s, col %s: %s",
                    row_idx,
                    llm_input_method_col,
                    llm_input_method,
                )
                ws.cell(row_idx, llm_input_method_col, llm_input_method)

                if not llm_chunked:
                    logger.info("  Single-pass summary (no chunking needed)")
            else:
                logger.warning("  ✗ LLM summary returned None")

            processed += 1

            # Save after each document to ensure data is persisted
            wb.save(self.metadata_path)
            logger.info("  💾 Document saved to Excel")

            # Also save periodically (redundant but safe)
            if processed % 5 == 0:
                logger.info("  💾 Checkpoint: %s documents processed", processed)

        # Final save
        wb.save(self.metadata_path)

        logger.info("\n" + "=" * 60)
        logger.info("✓ SUMMARIZATION COMPLETE")
        logger.info("=" * 60)
        logger.info("Processed: %s", processed)
        logger.info("Skipped: %s", skipped)


def main():
    """
    Main function to run the summarization pipeline from command line.

    Parses command-line arguments and executes the summarization process.
    """
    parser = argparse.ArgumentParser(
        description="Generate summaries for parsed documents"
    )
    parser.add_argument(
        "--metadata",
        default="./data/pdf_metadata.xlsx",
        help="Path to metadata Excel file (default: ./data/pdf_metadata.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Force re-summarization of all documents",
    )

    args = parser.parse_args()

    # Check if metadata file exists
    if not Path(args.metadata).exists():
        logger.error("Metadata file not found: %s", args.metadata)
        sys.exit(1)

    # Create summarizer
    summarizer = DocumentSummarizer(metadata_path=args.metadata)

    # Run summarization (LLM always enabled)
    summarizer.summarize_all(force_summarize=args.force)


if __name__ == "__main__":
    main()
